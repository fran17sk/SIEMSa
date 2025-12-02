from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.shortcuts import render, redirect
from django.views.generic import ListView
from django.contrib import messages
from .forms import ContratosForm
from django.http import JsonResponse
from .models import *
from .models_catastro import *
from .models_simsa import *
from django.utils.crypto import get_random_string
from email.mime.image import MIMEImage
from django.core.mail import EmailMultiAlternatives
from django.core.mail import EmailMessage
import json
import os
from django.db.models import OuterRef, Subquery
import json
import uuid
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models_catastro import Contratos
from datetime import datetime
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseForbidden
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string
from django.core.paginator import Paginator
from django.test import RequestFactory
from urllib.parse import urlencode
from django.db.models import Sum, Prefetch
from django.db.models.functions import ExtractYear
import geopandas as gpd
from .models import *
from .forms import MineralForm
from django.db.models import Sum, Q, Prefetch
from django.core.paginator import Paginator
from django.shortcuts import render
from django.core.mail import send_mail

from datetime import datetime
from collections import defaultdict
import json
import locale
import tempfile
import io
from django.db.models import F
from django.utils import timezone

# ReportLab - PDF
from reportlab.lib.pagesizes import letter, landscape, A4, A3
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Spacer, Image,
    Paragraph
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.pdfgen import canvas
import xlsxwriter

# Matplotlib - Gráficos
import matplotlib.pyplot as plt
from matplotlib.backends.backend_agg import FigureCanvasAgg as FigureCanvas

# OpenPyXL - Excel
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.conf import settings
from django.views.decorators.cache import never_cache

from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import User, Group
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test

import unicodedata
from difflib import get_close_matches
from django.http import JsonResponse
from django.views.decorators.http import require_GET
import pandas as pd
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.urls import reverse_lazy
from django.contrib.auth.views import PasswordChangeView
import random
import string
from django.contrib.auth.hashers import make_password
from django.db import transaction
import secrets
from django.views.generic import CreateView
from .models import PerfilUsuario, OrganismoUsuario
from datetime import date, timedelta

from django.db import connections
from django.db.utils import OperationalError
from django.core import serializers
from django.db.models import Subquery, OuterRef, Value
from django.db.models.functions import Cast
from django.db.models import UUIDField

from django.http import HttpResponse
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image
from django.templatetags.static import static
from django.db.models import OuterRef, Exists
import uuid
from openpyxl.utils import get_column_letter
import csv
from django.db import connection
from unidecode import unidecode

def group_required(group_names):
    def in_groups(u):
        if u.is_authenticated:
            if u.is_superuser:
                return True  # siempre permitir al superusuario
            if isinstance(group_names, str):
                groups = [group_names]
            else:
                groups = group_names
            return u.groups.filter(name__in=groups).exists()
        return False
    return user_passes_test(in_groups)

def formatear_fob(valor):
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


class CambioPasswordObligatorioView(PasswordChangeView):
    template_name = 'auth/obligatorio_password_change.html'
    success_url = reverse_lazy('home')  # o donde desees redirigir

    def form_valid(self, form):
        response = super().form_valid(form)
        perfil = self.request.user.perfilusuario
        perfil.debe_cambiar_password = False
        perfil.save()
        return response
    
@never_cache
def custom_login(request):
    if request.user.is_authenticated:
        logout(request)

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)

        if user:
            login(request, user)

            try:
                perfil = user.perfilusuario
                if perfil.debe_cambiar_password:
                    return JsonResponse({
                        "success": True,
                        "redirect_url": "/cambiar-password/"
                    })
            except PerfilUsuario.DoesNotExist:
                pass  # O manejar el caso si el perfil no existe

            return JsonResponse({
                "success": True,
                "redirect_url": settings.LOGIN_REDIRECT_URL
            })

        else:
            return JsonResponse({
                "success": False,
                "error": "Usuario o contraseña incorrectos"
            }, status=401)

    return render(request, 'auth/login.html')

def custom_logout(request):
    logout(request)
    return redirect('login')

@login_required(login_url='/login/')
def home(request):
    # tu código aquí
    return render(request, 'base.html')

@login_required(login_url='/login/')
def paises(request):
    paises = Pais.objects.all().order_by('nom_pais')  # ordenado alfabéticamente
    return render(request, 'exportaciones/Paises.html', {'paises': paises})


@login_required(login_url='/login/')
def productores_min(request):
    productores = ProdMinero.objects.all().order_by('nom_productor_min')
    return render(request, 'exportaciones/productores.html', {'productores': productores})



@login_required(login_url='/login/')
def minerales_list(request):
    minerales = Mineral.objects.all().order_by('-id_min')
    return render(request, 'exportaciones/minerales.html', {'minerales': minerales})

def mineral_create(request):
    if request.method == "POST":
        form = MineralForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('minerales_list')
    else:
        form = MineralForm()
    return render(request, 'exportaciones/minerales_create_modal.html', {'form': form})

def mineral_delete(request, pk):
    mineral = get_object_or_404(Mineral, pk=pk)
    if request.method == "POST":
        mineral.delete()
        return redirect('minerales_list')
    return render(request, 'exportaciones/minerales_delete_modal.html', {'mineral': mineral})



@login_required(login_url='/login/')
def new_exportacion(request,exportacion_id=None):

    if request.user.groups.filter(name='Lector').exists():
        return redirect('exportaciones')

    empresas = ProdMinero.objects.all().order_by('nom_productor_min')
    minerales = Mineral.objects.all().order_by('nom_min')
    paises = Pais.objects.all().order_by('nom_pais')

    if request.method == 'POST':
        try:
            data = request.POST
            detalles = json.loads(request.POST.get('detalles'))

            exportacion = Exportacion.objects.create(
                Num_Exped1=data.get('expediente'),
                fecha_export=data.get('fecha_exportacion'),
                id_productor_min_id=int(data.get('empresa')),
                id_pais_id=int(data.get('pais')),
                fecha_present_export=data.get('fecha_presentacion'),
                pedido_comercial_export=data.get('pedido_comercial'),
                Estado_anulacion=bool(data.get('anulacion')),
            )

            for det in detalles:
                MinExport.objects.create(
                    id_export=exportacion,
                    id_min_id=int(det['mineral_id']),
                    Tn_min_export=det['toneladas'],
                    FOB_min_export=det['valor_fob']
                )

            return JsonResponse({
                'status': 'success',
                'id_exportacion': exportacion.id_export
            })

        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            })
    ultima_exportacion = Exportacion.objects.order_by('-id_export').first()
    if not ultima_exportacion:
        id_ultima_exportacion = 0 
    else:
        id_ultima_exportacion = ultima_exportacion.id_export

    return render(request, 'exportaciones/new_exportacion.html', {
        'empresas': empresas,
        'minerales': minerales,
        'paises': paises,
        'numero_exportacion': id_ultima_exportacion + 1,
    })

def edit_exportacion(request, id_export):
    exportacion = get_object_or_404(Exportacion, id_export=id_export)
    detalles = MinExport.objects.filter(id_export=exportacion.id_export)

    if request.method == 'POST':
        # Actualizar los campos principales
        data = json.loads(request.body)
        print(data)

        declaracion = data.get('declaracion', {})

        # Guardar datos de la declaración
        exportacion.Num_Exped1 = declaracion.get('expediente')
        exportacion.fecha_export = declaracion.get('fecha_exportacion')
        exportacion.fecha_present_export = declaracion.get('fecha_presentacion')
        empresa = get_object_or_404(ProdMinero,nom_productor_min = declaracion.get('empresa'))
        exportacion.id_productor_min = empresa
        pais = get_object_or_404(Pais,nom_pais=declaracion.get('pais'))
        exportacion.id_pais = pais
        exportacion.pedido_comercial_export = declaracion.get('pedido_comercial')
        print(data['declaracion'].get('anulacion'))
        if data['declaracion'].get('anulacion'):
            exportacion.Estado_anulacion = 1
        else:
            exportacion.Estado_anulacion = 0

        exportacion.save()

        # Agregar nuevos detalles
        nuevos = data.get("nuevos", [])
        for detalle in nuevos:
            mineral_id = detalle.get("mineral_id")
            toneladas = detalle.get("toneladas")
            valor_fob = detalle.get("valor_fob")

            if mineral_id and toneladas and valor_fob:
                mineral = get_object_or_404(Mineral, id_min=mineral_id)
                MinExport.objects.create(
                    id_export=exportacion,
                    id_min=mineral,
                    Tn_min_export=toneladas,
                    FOB_min_export=valor_fob
                )

        

        # Editar detalles existentes
        editados = data.get("editados", [])
        for detalle in editados:
            detalle_id = detalle.get("id")
            toneladas = detalle.get("toneladas")
            valor_fob = detalle.get("valor_fob")

            if detalle_id and toneladas and valor_fob:
                det = get_object_or_404(MinExport, id_min_export=detalle_id, id_export=exportacion)
                det.Tn_min_export = toneladas
                det.FOB_min_export = valor_fob
                det.save()
        # Eliminar detalles por ID
        eliminados = data.get("eliminados", [])
        if eliminados:
            MinExport.objects.filter(id_min_export__in=eliminados, id_export=exportacion).delete()
        return JsonResponse({"message": "Exportación guardada correctamente."}, status=200)

    context = {
        'exportacion': exportacion,
        'detalles': detalles,
        'empresas': ProdMinero.objects.all(),
        'paises': Pais.objects.all(),
        'minerales': Mineral.objects.all(),
    }
    return render(request, 'exportaciones/edit_exportacion.html', context)

def exportacion_list(request):
    search = request.GET.get("search", "").strip()

    exportaciones_base = Exportacion.objects.select_related(
        'id_productor_min', 'id_pais'
    ).prefetch_related(
        Prefetch(
            'min_exports',
            queryset=MinExport.objects.select_related('id_min')
        )
    ).annotate(
        total_tn=Sum('min_exports__Tn_min_export'),
        total_fob=Sum('min_exports__FOB_min_export')
    ).order_by('-id_export')

    if search:
        exportaciones_base = exportaciones_base.filter(
            Q(id_productor_min__nom_productor_min__icontains=search) |
            Q(pedido_comercial_export = search) |
            Q(id_export = search) |
            Q(Num_Exped1 = search)
        ).distinct()

    paginator = Paginator(exportaciones_base, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 👉 Si es AJAX devolvemos JSON
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        data = []
        for exp in page_obj:
            data.append({
                "id": exp.id_export,
                "codigo": f"{exp.Num_Exped1}/{exp.Num_Exped2}" if exp.Num_Exped2 else exp.Num_Exped1,
                "fecha": exp.fecha_export.strftime("%d/%m/%Y"),
                "empresa": exp.id_productor_min.nom_productor_min,
                "pais": exp.id_pais.nom_pais,
                "pedido_comercial" : exp.pedido_comercial_export,
                "estado": exp.Estado_anulacion,
                "minerales": [me.id_min.nom_min for me in exp.min_exports.all()],
                "total_tn": f"{exp.total_tn:.2f}" if exp.total_tn else "0.00",
                "total_fob": f"{exp.total_fob:.2f}" if exp.total_fob else "0.00",
            })
        return JsonResponse({"exportaciones": data})

    # 👉 Render normal
    return render(request, 'exportaciones/exportaciones.html', {
        'page_obj': page_obj,
        'search': search,
    })







@csrf_exempt  # Solo si no estás usando el token CSRF en el JS (opcionalmente reemplazable)
def registrar_mineral(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'success': False, 'error': 'El nombre es requerido.'})
        
        nombre_normalizado = nombre.strip().upper()

        if Mineral.objects.filter(nom_min__iexact=nombre_normalizado).exists():
            return JsonResponse({'success': False, 'error': 'Ya existe un mineral con ese nombre.'})
        
        mineral = Mineral.objects.create(nom_min=nombre_normalizado)
        return JsonResponse({'success': True, 'id_min': mineral.id_min, 'nombre': mineral.nom_min})

    return JsonResponse({'success': False, 'error': 'Método no permitido.'})

@csrf_exempt  # Solo si no estás usando el token CSRF en el JS (opcionalmente reemplazable)
def registrar_pais(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'success': False, 'error': 'El nombre es requerido.'})
        
        nombre_normalizado = nombre.strip().upper()

        if Pais.objects.filter(nom_pais__iexact=nombre_normalizado).exists():
            return JsonResponse({'success': False, 'error': 'Ya existe un pais con ese nombre.'})
        
        pais = Pais.objects.create(nom_pais=nombre_normalizado)
        return JsonResponse({"success": True, "id": pais.id_pais, "nombre": pais.nom_pais})


    return JsonResponse({'success': False, 'error': 'Método no permitido.'})

@csrf_exempt  # Solo si no estás usando el token CSRF en el JS (opcionalmente reemplazable)
def registrar_productor(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        if not nombre:
            return JsonResponse({'success': False, 'error': 'El nombre es requerido.'})
        
        nombre_normalizado = nombre.strip().upper()

        if ProdMinero.objects.filter(nom_productor_min__iexact=nombre_normalizado).exists():
            return JsonResponse({'success': False, 'error': 'Ya existe una Empresa con ese nombre.'})
        
        productor = ProdMinero.objects.create(nom_productor_min=nombre_normalizado)
        return JsonResponse({"success": True,"id": productor.id_productor_min, "nombre": productor.nom_productor_min})

    return JsonResponse({'success': False, 'error': 'Método no permitido.'})





@csrf_exempt
def editar_mineral(request):
    if request.method == "POST":
        data = json.loads(request.body)
        mineral_id = data.get("id")
        nuevo_nombre = data.get("nombre", "").strip().upper()

        if not nuevo_nombre:
            return JsonResponse({"success": False, "error": "El nombre no puede estar vacío."})

        if Mineral.objects.filter(nom_min=nuevo_nombre).exclude(id_min=mineral_id).exists():
            return JsonResponse({"success": False, "error": "Ya existe un mineral con ese nombre."})

        try:
            mineral = Mineral.objects.get(id_min=mineral_id)
            mineral.nom_min = nuevo_nombre
            mineral.save()
            return JsonResponse({"success": True, "nombre": nuevo_nombre})
        except Mineral.DoesNotExist:
            return JsonResponse({"success": False, "error": "El mineral no existe."})
        

@csrf_exempt
def editar_pais(request):
    if request.method == "POST":
        data = json.loads(request.body)
        pais_id = data.get("id")
        nuevo_nombre = data.get("nombre", "").strip().upper()

        if not nuevo_nombre:
            return JsonResponse({"success": False, "error": "El nombre no puede estar vacío."})

        if Pais.objects.filter(nom_pais=nuevo_nombre).exclude(id_pais=pais_id).exists():
            return JsonResponse({"success": False, "error": "Ya existe un Pais con ese nombre."})

        try:
            pais = Pais.objects.get(id_pais=pais_id)
            pais.nom_pais = nuevo_nombre
            pais.save()
            return JsonResponse({"success": True, "nombre": nuevo_nombre})
        except Mineral.DoesNotExist:
            return JsonResponse({"success": False, "error": "El Pais no existe."})
        

@csrf_exempt
def editar_productor(request):
    if request.method == "POST":
        data = json.loads(request.body)
        productor_id = data.get("id")
        nuevo_nombre = data.get("nombre", "").strip().upper()

        if not nuevo_nombre:
            return JsonResponse({"success": False, "error": "El nombre no puede estar vacío."})

        if ProdMinero.objects.filter(nom_productor_min=nuevo_nombre).exclude(id_productor_min=productor_id).exists():
            return JsonResponse({"success": False, "error": "Ya existe una Empresa con ese nombre."})

        try:
            productor = ProdMinero.objects.get(id_productor_min=productor_id)
            productor.nom_productor_min = nuevo_nombre
            productor.save()
            return JsonResponse({"success": True, "nombre": nuevo_nombre})
        except Mineral.DoesNotExist:
            return JsonResponse({"success": False, "error": "La empresa no existe."})
        
@csrf_exempt
def guardar_exportacion(request, id_export):
    try:
        if request.method == 'POST':
            data = json.loads(request.body)

            declaracion = data.get('declaracion', {})

            exportacion = get_object_or_404(Exportacion, id=id_export)

            # Guardar datos de la declaración
            exportacion.Num_Exped1 = declaracion.get('expediente')
            exportacion.fecha_export = declaracion.get('fecha_exportacion')
            exportacion.fecha_present_export = declaracion.get('fecha_presentacion')
            exportacion.id_productor_min = declaracion.get('empresa')
            exportacion.id_pais = declaracion.get('pais')
            exportacion.pedido_comercial_export = declaracion.get('pedido_comercial')
            exportacion.detalles = declaracion.get('detalles', '')
            #exportacion.save()

            # Agregar nuevos detalles
            nuevos = data.get("nuevos", [])
            for detalle in nuevos:
                mineral_id = detalle.get("mineral_id")
                toneladas = detalle.get("toneladas")
                valor_fob = detalle.get("valor_fob")

                if mineral_id and toneladas and valor_fob:
                    mineral = get_object_or_404(Mineral, id=mineral_id)
                    MinExport.objects.create(
                        exportacion=exportacion,
                        mineral=mineral,
                        toneladas=toneladas,
                        valor_fob=valor_fob
                    )

            # Eliminar detalles por ID
            eliminados = data.get("eliminados", [])
            if eliminados:
                MinExport.objects.filter(id__in=eliminados, exportacion=exportacion).delete()

            # Editar detalles existentes
            editados = data.get("editados", [])
            for detalle in editados:
                detalle_id = detalle.get("id")
                toneladas = detalle.get("toneladas")
                valor_fob = detalle.get("valor_fob")

                if detalle_id and toneladas and valor_fob:
                    det = get_object_or_404(MinExport, id=detalle_id, exportacion=exportacion)
                    det.toneladas = toneladas
                    det.valor_fob = valor_fob
                    det.save()

            return JsonResponse({"message": "Exportación guardada correctamente."}, status=200)

        return JsonResponse({"message": "Método no permitido."}, status=405)

    except Exception as e:
        return JsonResponse({"message": f"Error al guardar: {str(e)}"}, status=400)
    





def dashboard_exportaciones(request):
    # Obtener años únicos
    anios = sorted(set(
        Exportacion.objects.dates('fecha_export', 'year')
    ), key=lambda x: x.year)

    # Estructura de datos: { año: { mineral: toneladas } }
    datos = defaultdict(dict)
    minerales_set = set()

    for anio in anios:
        resultados = MinExport.objects.filter(
            id_export__fecha_export__year=anio.year
        ).values(
            'id_min__nom_min'
        ).annotate(
            total=Sum('Tn_min_export')
        )

        for resultado in resultados:
            mineral = resultado['id_min__nom_min']
            toneladas = float(resultado['total'])
            datos[anio.year][mineral] = toneladas
            minerales_set.add(mineral)

    minerales = sorted(minerales_set)
    datos_por_anio = {
        str(anio.year): [datos[anio.year].get(mineral, 0) for mineral in minerales]
        for anio in anios
    }

    context = {
        'minerales': minerales,
        'datos_por_anio': datos_por_anio,
        'anios': [a.year for a in anios],
    }

    return render(request, 'exportaciones/graficos.html', context)


def dashboard(request):
    years = MinExport.objects.annotate(
        year=ExtractYear('id_export__fecha_export')
    ).values_list('year', flat=True).distinct().order_by('-year')
    minerales = Mineral.objects.all()
    context = {
        'years': years,
        'current_year': datetime.now().year,
        'total_empresas' : ProdMinero.objects.count(),
        'total_paises' : Pais.objects.filter(exportaciones__isnull=False).distinct().count(),
        'total_minerales' : Mineral.objects.count(),
        'total_exportaciones' : Exportacion.objects.filter(Estado_anulacion=False).count(),
        'minerales':minerales
    }
    return render(request, 'exportaciones/dashboard.html',context)


def datos_toneladas_exportadas(request):
    anios = sorted(set(Exportacion.objects.dates('fecha_export', 'year')), key=lambda x: x.year)
    datos = defaultdict(dict)
    minerales_set = set()

    for anio in anios:
        resultados = MinExport.objects.filter(
            id_export__fecha_export__year=anio.year
        ).values(
            'id_min__nom_min'
        ).annotate(
            total=Sum('Tn_min_export')
        )

        for resultado in resultados:
            mineral = resultado['id_min__nom_min']
            toneladas = float(resultado['total'])
            datos[anio.year][mineral] = toneladas
            minerales_set.add(mineral)

    minerales = sorted(minerales_set)

    # Gráfico de líneas: minerales en X, líneas por año
    datos_lineas = {
        str(anio.year): [datos[anio.year].get(mineral, 0) for mineral in minerales]
        for anio in anios
    }

    # Gráfico de barras: años en X, total por mineral
    datos_barras = {}
    for mineral in minerales:
        datos_barras[mineral] = [
            datos[anio.year].get(mineral, 0) for anio in anios
        ]

    # Totales por año
    total_por_anio = [
        sum(datos[anio.year].values()) for anio in anios
    ]

    return JsonResponse({
        'anios': [a.year for a in anios],
        'minerales': minerales,
        'datos_lineas': datos_lineas,
        'datos_barras': datos_barras,
        'totales': total_por_anio,
    })


def datos_fob_exportadas(request):
    anios = sorted(set(Exportacion.objects.dates('fecha_export', 'year')), key=lambda x: x.year)
    datos = defaultdict(dict)
    minerales_set = set()

    for anio in anios:
        resultados = MinExport.objects.filter(
            id_export__fecha_export__year=anio.year
        ).values(
            'id_min__nom_min'
        ).annotate(
            total=Sum('FOB_min_export')
        )

        for resultado in resultados:
            mineral = resultado['id_min__nom_min']
            toneladas = float(resultado['total'])
            datos[anio.year][mineral] = toneladas
            minerales_set.add(mineral)

    minerales = sorted(minerales_set)

    # Gráfico de líneas: minerales en X, líneas por año
    datos_lineas = {
        str(anio.year): [datos[anio.year].get(mineral, 0) for mineral in minerales]
        for anio in anios
    }

    # Gráfico de barras: años en X, total por mineral
    datos_barras = {}
    for mineral in minerales:
        datos_barras[mineral] = [
            datos[anio.year].get(mineral, 0) for anio in anios
        ]

    # Totales por año
    total_por_anio = [
        sum(datos[anio.year].values()) for anio in anios
    ]

    return JsonResponse({
        'anios': [a.year for a in anios],
        'minerales': minerales,
        'datos_lineas': datos_lineas,
        'datos_barras': datos_barras,
        'totales': total_por_anio,
    })


def exportar_toneladas(request, formato):
    # Obtener últimos 5 años
    anios = sorted(set(Exportacion.objects.dates('fecha_export', 'year')), key=lambda x: x.year)[-5:]

    datos = defaultdict(dict)
    minerales_set = set()

    for anio in anios:
        resultados = MinExport.objects.filter(
            id_export__fecha_export__year=anio.year
        ).values(
            'id_min__nom_min'
        ).annotate(
            total=Sum('Tn_min_export')
        )

        for resultado in resultados:
            mineral = resultado['id_min__nom_min']
            toneladas = float(resultado['total'])
            datos[anio.year][mineral] = toneladas
            minerales_set.add(mineral)

    minerales = sorted(minerales_set)

    # Crear tabla: primera fila con años + "Total"
    header = ["Mineral"] + [str(a.year) for a in anios] + ["Total"]
    datos_tabla = [header]

    # Filas: cada mineral con valores por año y total por mineral
    for mineral in minerales:
        fila = [mineral]
        total_min = 0
        for anio in anios:
            val = datos[anio.year].get(mineral, 0)
            fila.append(f"{val:.2f}")
            total_min += val
        fila.append(f"{total_min:.2f}")
        datos_tabla.append(fila)

    if formato == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30,leftMargin=30, topMargin=30,bottomMargin=18)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'title',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=20,
        )
        title = Paragraph("Reporte de Cantidades Exportadas en toneladas (últimos 5 años)", title_style)
        elements.append(title)

        # Ajustar anchos de columna: mineral más ancho, años y total iguales
        col_widths = [150] + [80]*len(anios) + [80]

        t = Table(datos_tabla, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4a90e2")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 24))

        # Funciones para gráficos idénticas a antes

        def crear_grafico_lineas():
            fig, ax = plt.subplots(figsize=(12, 5))
            for mineral in minerales:
                y = [datos[anio.year].get(mineral, 0) for anio in anios]
                ax.plot([a.year for a in anios], y, label=mineral, marker='o')
            ax.set_title('Toneladas por Mineral por Año')
            ax.set_xlabel('Año')
            ax.set_ylabel('Toneladas')
            ax.legend(loc='upper left', bbox_to_anchor=(1,1))
            ax.grid(True)
            fig.tight_layout()
            return fig

        def crear_grafico_barras():
            fig, ax = plt.subplots(figsize=(12, 5))
            total_por_anio = [sum(datos[anio.year].values()) for anio in anios]
            ax.bar([a.year for a in anios], total_por_anio, color='skyblue')
            ax.set_title('Toneladas Totales por Año')
            ax.set_xlabel('Año')
            ax.set_ylabel('Toneladas Totales')
            ax.grid(axis='y')
            fig.tight_layout()
            return fig

        for crear_grafico in [crear_grafico_lineas, crear_grafico_barras]:
            fig = crear_grafico()
            img_buffer = io.BytesIO()
            canvas = FigureCanvas(fig)
            canvas.print_png(img_buffer)
            plt.close(fig)
            img_buffer.seek(0)
            img = Image(img_buffer, width=720, height=300)
            elements.append(img)
            elements.append(Spacer(1, 20))

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="toneladas_exportadas.pdf"'
        response.write(pdf)
        return response
    elif formato == 'excel':
        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("Toneladas")

        bold = workbook.add_format({'bold': True, 'bg_color': '#4a90e2', 'color': 'white', 'align': 'center'})
        normal = workbook.add_format({'align': 'center'})

        # Escribir la tabla
        for i, fila in enumerate(datos_tabla):
            for j, celda in enumerate(fila):
                fmt = bold if i == 0 else normal
                worksheet.write(i, j, celda, fmt)

        worksheet.autofilter(0, 0, len(datos_tabla) - 1, len(header) - 1)
        worksheet.freeze_panes(1, 1)

        # Crear gráficos como imágenes
        def crear_grafico_lineas():
            fig, ax = plt.subplots(figsize=(8, 4))
            for mineral in minerales:
                y = [datos[anio.year].get(mineral, 0) for anio in anios]
                ax.plot([a.year for a in anios], y, label=mineral, marker='o')
            ax.set_title('Toneladas por Mineral por Año')
            ax.set_xlabel('Año')
            ax.set_ylabel('Toneladas')
            ax.legend(loc='upper left', bbox_to_anchor=(1,1))
            ax.grid(True)
            fig.tight_layout()
            return fig

        def crear_grafico_barras():
            fig, ax = plt.subplots(figsize=(8, 4))
            total_por_anio = [sum(datos[anio.year].values()) for anio in anios]
            ax.bar([a.year for a in anios], total_por_anio, color='skyblue')
            ax.set_title('Toneladas Totales por Año')
            ax.set_xlabel('Año')
            ax.set_ylabel('Toneladas Totales')
            ax.grid(axis='y')
            fig.tight_layout()
            return fig

        # Agregar hoja con gráficos
        for i, (crear_grafico, nombre_hoja) in enumerate([
            (crear_grafico_lineas, "Gráfico Líneas"),
            (crear_grafico_barras, "Gráfico Barras")
        ]):
            fig = crear_grafico()
            img_buffer = io.BytesIO()
            canvas = FigureCanvas(fig)
            canvas.print_png(img_buffer)
            plt.close(fig)
            img_buffer.seek(0)

            hoja = workbook.add_worksheet(nombre_hoja)
            hoja.insert_image('B2', f'grafico_{i}.png', {'image_data': img_buffer})

        workbook.close()
        output.seek(0)

        response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="toneladas_exportadas.xlsx"'
        return response

def exportar_fob(request, formato):
    # Obtener últimos 5 años
    anios = sorted(set(Exportacion.objects.dates('fecha_export', 'year')), key=lambda x: x.year)[-5:]

    datos = defaultdict(dict)
    minerales_set = set()

    for anio in anios:
        resultados = MinExport.objects.filter(
            id_export__fecha_export__year=anio.year
        ).values(
            'id_min__nom_min'
        ).annotate(
            total=Sum('FOB_min_export')
        )

        for resultado in resultados:
            mineral = resultado['id_min__nom_min']
            toneladas = float(resultado['total'])
            datos[anio.year][mineral] = toneladas
            minerales_set.add(mineral)

    minerales = sorted(minerales_set)

    # Crear tabla: primera fila con años + "Total"
    header = ["Mineral"] + [str(a.year) for a in anios] + ["Total"]
    datos_tabla = [header]

    # Filas: cada mineral con valores por año y total por mineral
    for mineral in minerales:
        fila = [mineral]
        total_min = 0
        for anio in anios:
            val = datos[anio.year].get(mineral, 0)
            fila.append(formatear_fob(val))
            total_min += val
        fila.append(formatear_fob(total_min))
        datos_tabla.append(fila)

    if formato == 'pdf':
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), rightMargin=30,leftMargin=30, topMargin=30,bottomMargin=18)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'title',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=20,
        )
        title = Paragraph("Reporte de Cantidades Exportadas en Valor FOB USD (últimos 5 años)", title_style)
        elements.append(title)

        # Ajustar anchos de columna: mineral más ancho, años y total iguales
        col_widths = [150] + [80]*len(anios) + [80]

        t = Table(datos_tabla, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4a90e2")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 1), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 24))

        # Funciones para gráficos idénticas a antes

        def crear_grafico_lineas():
            fig, ax = plt.subplots(figsize=(12, 5))
            for mineral in minerales:
                y = [datos[anio.year].get(mineral, 0) for anio in anios]
                ax.plot([a.year for a in anios], y, label=mineral, marker='o')
            ax.set_title('FOB (USD) por Mineral por Año')
            ax.set_xlabel('Año')
            ax.set_ylabel('Valor FOB')
            ax.legend(loc='upper left', bbox_to_anchor=(1,1))
            ax.grid(True)
            fig.tight_layout()
            return fig

        def crear_grafico_barras():
            fig, ax = plt.subplots(figsize=(12, 5))
            total_por_anio = [sum(datos[anio.year].values()) for anio in anios]
            ax.bar([a.year for a in anios], total_por_anio, color='skyblue')
            ax.set_title('Valor FOB Totales por Año')
            ax.set_xlabel('Año')
            ax.set_ylabel('FOB Totales')
            ax.grid(axis='y')
            fig.tight_layout()
            return fig

        for crear_grafico in [crear_grafico_lineas, crear_grafico_barras]:
            fig = crear_grafico()
            img_buffer = io.BytesIO()
            canvas = FigureCanvas(fig)
            canvas.print_png(img_buffer)
            plt.close(fig)
            img_buffer.seek(0)
            img = Image(img_buffer, width=720, height=300)
            elements.append(img)
            elements.append(Spacer(1, 20))

        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="toneladas_exportadas.pdf"'
        response.write(pdf)
        return response
    elif formato == 'excel':
        header = datos_tabla[0]

        output = io.BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet("FOB")

        bold = workbook.add_format({'bold': True, 'bg_color': '#4a90e2', 'color': 'white', 'align': 'center'})
        normal = workbook.add_format({'align': 'center'})

        # Escribir la tabla
        for i, fila in enumerate(datos_tabla):
            for j, celda in enumerate(fila):
                fmt = bold if i == 0 else normal
                worksheet.write(i, j, celda, fmt)

        worksheet.autofilter(0, 0, len(datos_tabla) - 1, len(header) - 1)
        worksheet.freeze_panes(1, 1)

        # Crear gráficos como imágenes
        def crear_grafico_lineas():
            fig, ax = plt.subplots(figsize=(8, 4))
            for mineral in minerales:
                y = [datos[anio.year].get(mineral, 0) for anio in anios]
                ax.plot([a.year for a in anios], y, label=mineral, marker='o')
            ax.set_title('FOB por Mineral por Año')
            ax.set_xlabel('Año')
            ax.set_ylabel('FOB (USD)')
            ax.legend(loc='upper left', bbox_to_anchor=(1,1))
            ax.grid(True)
            fig.tight_layout()
            return fig

        def crear_grafico_barras():
            fig, ax = plt.subplots(figsize=(8, 4))
            total_por_anio = [sum(datos[anio.year].values()) for anio in anios]
            ax.bar([a.year for a in anios], total_por_anio, color='skyblue')
            ax.set_title('FOB Total por Año')
            ax.set_xlabel('Año')
            ax.set_ylabel('FOB Total (USD)')
            ax.grid(axis='y')
            fig.tight_layout()
            return fig

        # Agregar hoja con gráficos
        for i, (crear_grafico, nombre_hoja) in enumerate([
            (crear_grafico_lineas, "Gráfico Líneas"),
            (crear_grafico_barras, "Gráfico Barras")
        ]):
            fig = crear_grafico()
            img_buffer = io.BytesIO()
            canvas = FigureCanvas(fig)
            canvas.print_png(img_buffer)
            plt.close(fig)
            img_buffer.seek(0)

            hoja = workbook.add_worksheet(nombre_hoja)
            hoja.insert_image('B2', f'grafico_{i}.png', {'image_data': img_buffer})

        workbook.close()
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="fob_exportado.xlsx"'
        return response

def top_productores_exportacion(request):
    year = request.GET.get('year')

    queryset = MinExport.objects.all()
    if year:
        queryset = queryset.filter(id_export__fecha_export__year=year)

    resultados = (
        queryset
        .values('id_export__id_productor_min__nom_productor_min')
        .annotate(total=Sum('Tn_min_export'))
        .order_by('-total')[:5]
    )

    productores = [r['id_export__id_productor_min__nom_productor_min'] for r in resultados]
    valores = [round(r['total'] or 0, 2) for r in resultados]

    return JsonResponse({
        'empresas': productores,
        'valores': valores,
    })

def top_productores_valor_fob(request):
    year = request.GET.get('year')

    queryset = MinExport.objects.all()
    if year:
        queryset = queryset.filter(id_export__fecha_export__year=year)

    resultados = (
        queryset
        .values('id_export__id_productor_min__nom_productor_min')
        .annotate(total_fob=Sum('FOB_min_export'))
        .order_by('-total_fob')[:5]
    )

    empresas = [r['id_export__id_productor_min__nom_productor_min'] for r in resultados]
    valores = [round(r['total_fob'] or 0, 2) for r in resultados]

    return JsonResponse({
        'empresas': empresas,
        'valores': valores,
    })

def exportar_excel_top_productores(request, year):
    year = int(year)
    wb = Workbook()
    ws = wb.active
    ws.title = f'Top Productores {year}'

    # ============================
    # TÍTULO
    # ============================
    titulo = f"Reporte Top 5 Productores Mineros - {year}"
    ws.merge_cells('A1:D1')
    cell = ws['A1']
    cell.value = titulo
    cell.font = Font(size=14, bold=True)
    cell.alignment = Alignment(horizontal='center')

    ws.append([])

    # ============================
    # TOP 5 POR TONELADAS
    # ============================
    toneladas_data = (
        MinExport.objects
        .filter(id_export__fecha_export__year=year)
        .values('id_export__id_productor_min__id_productor_min',
                'id_export__id_productor_min__nom_productor_min')
        .annotate(total_tn=Sum('Tn_min_export'))
        .order_by('-total_tn')[:5]
    )

    ws.append(["Top 5 Productores por Toneladas", "Toneladas Exportadas"])
    ws.append([])

    for prod in toneladas_data:
        productor_nombre = prod['id_export__id_productor_min__nom_productor_min']
        productor_id = prod['id_export__id_productor_min__id_productor_min']
        total_tn = round(prod['total_tn'], 2)

        ws.append([productor_nombre, total_tn])
        ws.append(["Mineral", "Toneladas", "FOB (USD)"])

        # Detalle por mineral
        minerales = (
            MinExport.objects
            .filter(id_export__id_productor_min__id_productor_min=productor_id,
                    id_export__fecha_export__year=year)
            .values("id_min__nom_min")
            .annotate(
                total_tn=Sum("Tn_min_export"),
                total_fob=Sum("FOB_min_export")
            )
            .order_by("-total_tn")
        )

        for m in minerales:
            ws.append([
                m["id_min__nom_min"],
                float(m["total_tn"]),
                float(m["total_fob"]),
            ])

        ws.append([])

    ws.append([])
    ws.append([])

    # ============================
    # TOP 5 POR VALOR FOB
    # ============================
    valor_data = (
        MinExport.objects
        .filter(id_export__fecha_export__year=year)
        .values('id_export__id_productor_min__id_productor_min',
                'id_export__id_productor_min__nom_productor_min')
        .annotate(total_fob=Sum('FOB_min_export'))
        .order_by('-total_fob')[:5]
    )

    ws.append(["Top 5 Productores por Valor FOB (USD)", "Valor Total Exportado"])
    ws.append([])

    for prod in valor_data:
        productor_nombre = prod['id_export__id_productor_min__nom_productor_min']
        productor_id = prod['id_export__id_productor_min__id_productor_min']
        total_fob = round(prod['total_fob'], 2)

        ws.append([productor_nombre, total_fob])
        ws.append(["Mineral", "Toneladas", "FOB (USD)"])

        # Detalle por mineral
        minerales = (
            MinExport.objects
            .filter(id_export__id_productor_min__id_productor_min=productor_id,
                    id_export__fecha_export__year=year)
            .values("id_min__nom_min")
            .annotate(
                total_tn=Sum("Tn_min_export"),
                total_fob=Sum("FOB_min_export")
            )
            .order_by("-total_fob")
        )

        for m in minerales:
            ws.append([
                m["id_min__nom_min"],
                float(m["total_tn"]),
                float(m["total_fob"]),
            ])

        ws.append([])

    # ============================
    # RESPUESTA
    # ============================
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'top_productores_detallado_{year}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def exportar_pdf_top_productores(request, year):
    year = int(year)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph(f"<b>Reporte Top 5 Productores Mineros ({year})</b>", styles['Title']))
    elements.append(Spacer(1, 12))

    # Toneladas
    toneladas_data = (
        MinExport.objects.filter(id_export__fecha_export__year=year)
        .values('id_export__id_productor_min__nom_productor_min')
        .annotate(total=Sum('Tn_min_export'))
        .order_by('-total')[:5]
    )
    elements.append(Paragraph("Top 5 por Toneladas Exportadas (en Tn)", styles['Heading2']))
    tabla_toneladas = [["Empresa", "Toneladas"]]
    total_ton = sum(r['total'] for r in toneladas_data)
    for r in toneladas_data:
        nombre = r['id_export__id_productor_min__nom_productor_min']
        valor = round(r['total'], 2)
        porcentaje = f"{(valor / total_ton) * 100:.2f} %"
        tabla_toneladas.append([nombre, f"{valor} ({porcentaje})"])
    elements.append(Table(tabla_toneladas))
    elements.append(Spacer(1, 20))

    # Valor FOB
    valor_data = (
        MinExport.objects.filter(id_export__fecha_export__year=year)
        .values('id_export__id_productor_min__nom_productor_min')
        .annotate(total=Sum('FOB_min_export'))
        .order_by('-total')[:5]
    )
    elements.append(Paragraph("Top 5 por Valor FOB Exportado (USD)", styles['Heading2']))
    tabla_valor = [["Empresa", "Valor FOB"]]
    total_val = sum(r['total'] for r in valor_data)
    for r in valor_data:
        nombre = r['id_export__id_productor_min__nom_productor_min']
        valor = round(r['total'], 2)
        porcentaje = f"{(valor / total_val) * 100:.2f} %"
        tabla_valor.append([nombre, f"${valor} ({porcentaje})"])
    elements.append(Table(tabla_valor))

    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')



def export_data_api(request):
    year = request.GET.get('year')
    mineral_id = request.GET.get('mineral')
    PAISES_MAP = {
            "ARGENTINA": "Argentina",
            "BRASIL": "Brazil",
            "BOLIVIA": "Bolivia",
            "CHILE": "Chile",
            "MALASIA": "Malaysia",
            "REINO UNIDO": "United Kingdom",
            "BANGLADESH": "Bangladesh",
            "ESTADOS UNIDOS": "United States of America",
            "INDIA": "India",
            "AUSTRALIA": "Australia",
            "PAKISTAN": "Pakistan",
            "CHINA": "China",
            "HOLANDA": "Netherlands",
            "POLONIA": "Poland",
            "BELGICA": "Belgium",
            "NUEVA ZELANDA": "New Zealand",
            "SUDAFRICA": "South Africa",
            "FRANCIA": "France",
            "JAPON": "Japan",
            "ITALIA": "Italy",
            "ESPAÑA": "Spain",
            "ALEMANIA": "Germany",
            "INDONESIA": "Indonesia",
            "COSTA RICA": "Costa Rica",
            "CANADA": "Canada",
            "MEXICO": "Mexico",
            "EGIPTO": "Egypt",
            "NICARAGUA": "Nicaragua",
            "URUGUAY": "Uruguay",
            "PARAGUAY": "Paraguay",
            "ARABIA SAUDITA": "Saudi Arabia",
            "COLOMBIA": "Colombia",
            "COREA": "South Korea",
            "KOREA": "South Korea",
            "ECUADOR": "Ecuador",
            "GUATEMALA": "Guatemala",
            "IRAN": "Iran",
            "KENIA": "Kenya",
            "MARRUECOS": "Morocco",
            "PUERTO RICO": "Puerto Rico",
            "TAIWAN": "Taiwan",
            "TAILANDIA": "Thailand",
            "THAILANDIA": "Thailand",
            "VENEZUELA": "Venezuela",
            "PERU": "Peru",
            "PANAMA": "Panama",
            "MONTENEGRO": "Montenegro",
            "TURQUIA": "Turkey",
            "RUSIA": "Russia",
            "MADAGASCAR": "Madagascar",
            "EMIRATOS ARABES UNID": "United Arab Emirates",
            "HONDURAS": "Honduras",
            "REP. DOMINICANA": "Dominican Republic",
            "REPUBLICA DOMINICANA": "Dominican Republic",
            "VIETNAM": "Vietnam",
            "NEPAL": "Nepal",
            "TUNEZ": "Tunisia",
            "PAPUA NUEVA GUINEA": "Papua New Guinea",
            "NIGERIA": "Nigeria",
            "BIRMANIA": "Myanmar",
            "ARGELIA": "Algeria",
            "EL SALVADOR": "El Salvador",
            "SRI LANKA": "Sri Lanka",
            "INGLATERRA": "United Kingdom",
            "IRLANDA DEL SUR": "Ireland",
            "HUNGRIA": "Hungary",
            "SENEGAL": "Senegal",
            "SIRIA": "Syria",
            "HONG KONG": "Hong Kong",
            "DINAMARCA": "Denmark",
            "IRLANDA": "Ireland",
            "SINGAPUR": "Singapore",
            "RUMANIA": "Romania",
            "FILIPINAS": "Philippines",
            "ISRAEL": "Israel",
            "JORDANIA":"Jordan"
        }

    qs = MinExport.objects.all()

    if year:
        qs = qs.filter(id_export__fecha_export__year=year)

    if mineral_id:
        qs = qs.filter(id_min=mineral_id)

    data = (
        qs
        .values('id_export__id_pais__nom_pais', 'id_export__id_pais__id_pais')
        .annotate(
            total_toneladas=Sum('Tn_min_export'),
            total_fob=Sum('FOB_min_export')
        )
    )

    response_data = []
    for entry in data:
        nom_pais = entry['id_export__id_pais__nom_pais'].upper()
        nom_normalizado = PAISES_MAP.get(nom_pais, entry['id_export__id_pais__nom_pais'])
        response_data.append({
            'pais': nom_normalizado,
            'id_pais': entry['id_export__id_pais__id_pais'],
            'toneladas': float(entry['total_toneladas']),
            'fob': float(entry['total_fob']),
            'year': year
        })

    return JsonResponse(response_data, safe=False)

def exportar_pdf(request):
    year = request.GET.get("year")
    mineral_nombre = request.GET.get("mineral_nombre", "")

    # Definir años a mostrar
    if year and year.lower() != "todos":
        years = [int(year)]
    else:
        actual_year = datetime.now().year
        years = list(range(actual_year - 4, actual_year + 1))

    # Obtener datos desde la API original
    parsed_data = []
    factory = RequestFactory()

    if not year or year.lower() == "todos":
        for y in years:
            # Clonar GET y asignar el año correcto
            params = request.GET.copy()
            params["year"] = str(y)

            query_string = urlencode(params)
            url = f"{request.path}?{query_string}"


            fake_req = factory.get(url)
            fake_req.user = request.user

            response = export_data_api(fake_req)
            year_data = json.loads(response.content)

            for item in year_data:
                if "year" not in item or not item["year"]:
                    continue
                parsed_data.append(item)
    else:
        response = export_data_api(request)
        parsed_data = json.loads(response.content)
    # Organizar datos
    report_data = {}
    for item in parsed_data:
        if "year" not in item or not item["year"]:
            continue

        anio = int(item["year"])
        
        # Filtro explícito por año consultado (cuando no es "todos")
        if year and year.lower() != "todos" and anio != int(year):
            continue

        pais = item["pais"]
        toneladas = float(item.get("toneladas", 0))
        fob = float(item.get("fob", 0))

        if pais not in report_data:
            report_data[pais] = {}
        report_data[pais][anio] = {'toneladas': toneladas, 'fob': fob}

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A3),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20,
    )
    elements = []
    styles = getSampleStyleSheet()

    title = "Reporte de exportaciones por países, anuales"
    if year and year.lower() != "todos":
        title += f" - Año {year}"
    if mineral_nombre:
        title += f" - Mineral: {mineral_nombre}"

    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Spacer(1, 12))

    # Construir encabezados
    header = ["País"]
    for y in years:
        header.append(f"Tn {y}")
        header.append(f"FOB {y}")
    header.append("Totales Tn")
    header.append("Totales FOB")

    table_data = [header]

    total_por_columna = {y: {'toneladas': 0, 'fob': 0} for y in years}
    total_general_tn = 0
    total_general_fob = 0

    for pais, anios_data in report_data.items():
        fila = [pais]
        suma_tn_fila = 0
        suma_fob_fila = 0
        for y in years:
            tn = anios_data.get(y, {}).get('toneladas', 0)
            fob = anios_data.get(y, {}).get('fob', 0)
            fila.append(f"{tn:,.2f}")
            fila.append(f"{fob:,.2f}")

            suma_tn_fila += tn
            suma_fob_fila += fob
            total_por_columna[y]['toneladas'] += tn
            total_por_columna[y]['fob'] += fob

        fila.append(f"{suma_tn_fila:,.2f}")
        fila.append(f"{suma_fob_fila:,.2f}")

        total_general_tn += suma_tn_fila
        total_general_fob += suma_fob_fila

        table_data.append(fila)

    # Fila de totales
    fila_totales = ["Totales"]
    for y in years:
        fila_totales.append(f"{total_por_columna[y]['toneladas']:,.2f}")
        fila_totales.append(f"{total_por_columna[y]['fob']:,.2f}")
    fila_totales.append(f"{total_general_tn:,.2f}")
    fila_totales.append(f"{total_general_fob:,.2f}")

    table_data.append(fila_totales)

    col_widths = [120] + [80] * (2 * len(years)) + [90, 90]
    t = Table(table_data, repeatRows=1, colWidths=col_widths)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#d3d3d3")),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor("#e0e0e0")),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('ALIGN', (1, 1), (-1, -2), 'RIGHT'),
        ('ALIGN', (1, -1), (-1, -1), 'RIGHT'),
        ('ALIGN', (0, 0), (0, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ])
    t.setStyle(style)
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(
        buffer,
        content_type='application/pdf',
        headers={'Content-Disposition': 'attachment; filename="reporte_exportaciones.pdf"'}
    )


def exportar_excel(request):
    year = request.GET.get("anio")
    mineral_id = request.GET.get("mineral")
    mineral_nombre = request.GET.get("mineral_nombre", "")

    qs = MinExport.objects.all()
    if mineral_id:
        qs = qs.filter(id_min=mineral_id)

    if not year or year.lower() == "todos":
        current_year = datetime.now().year
        years = list(range(current_year - 4, current_year + 1))
        qs = qs.filter(id_export__fecha_export__year__in=years)
    else:
        years = [int(year)]
        qs = qs.filter(id_export__fecha_export__year=year)

    data = (
        qs
        .values('id_export__id_pais__nom_pais', 'id_export__fecha_export__year')
        .annotate(
            total_toneladas=Sum('Tn_min_export'),
            total_fob=Sum('FOB_min_export')
        )
        .order_by('id_export__id_pais__nom_pais', 'id_export__fecha_export__year')
    )

    PAISES_MAP = {
        # ... tu diccionario de países ...
    }

    report_data = {}
    for entry in data:
        pais_raw = entry['id_export__id_pais__nom_pais'].upper()
        pais = PAISES_MAP.get(pais_raw, entry['id_export__id_pais__nom_pais'])
        anio = entry['id_export__fecha_export__year']
        tonel = entry['total_toneladas'] or 0
        fob = entry['total_fob'] or 0

        if pais not in report_data:
            report_data[pais] = {}
        report_data[pais][anio] = {
            'toneladas': tonel,
            'fob': fob
        }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exportaciones"

    titulo = "Reporte de exportaciones por países, anuales"
    if year and year.lower() != "todos":
        titulo += f" - Año {year}"
    if mineral_nombre:
        titulo += f" - Mineral: {mineral_nombre}"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(years)*2 + 1)
    ws.cell(row=1, column=1).value = titulo
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    # Encabezados
    ws.cell(row=2, column=1).value = "País"
    ws.cell(row=2, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)

    col = 2
    for y in years:
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col+1)
        ws.cell(row=2, column=col).value = f"Año {y}"
        ws.cell(row=2, column=col).font = Font(bold=True)
        ws.cell(row=2, column=col).alignment = Alignment(horizontal="center")

        ws.cell(row=3, column=col).value = "Tn"
        ws.cell(row=3, column=col).font = Font(bold=True)
        ws.cell(row=3, column=col+1).value = "FOB (USD)"
        ws.cell(row=3, column=col+1).font = Font(bold=True)

        col += 2

    # Columna extra para totales por fila
    ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)
    ws.cell(row=2, column=col).value = "Totales"
    ws.cell(row=2, column=col).font = Font(bold=True)
    ws.cell(row=2, column=col).alignment = Alignment(horizontal="center")

    # Datos y totales por fila
    fila = 4
    total_por_columna = {y: {'toneladas': 0, 'fob': 0} for y in years}
    total_general_toneladas = 0
    total_general_fob = 0

    for pais, anios_data in report_data.items():
        ws.cell(row=fila, column=1).value = pais
        suma_tn_fila = 0
        suma_fob_fila = 0
        col = 2
        for y in years:
            toneladas = anios_data.get(y, {}).get('toneladas', 0)
            fob = anios_data.get(y, {}).get('fob', 0)

            ws.cell(row=fila, column=col).value = round(toneladas, 2)
            ws.cell(row=fila, column=col+1).value = round(fob, 2)

            suma_tn_fila += toneladas
            suma_fob_fila += fob

            total_por_columna[y]['toneladas'] += toneladas
            total_por_columna[y]['fob'] += fob

            col += 2

        # Total fila: mostrar suma tn + fob juntos (puede separar si querés)
        ws.cell(row=fila, column=col).value = f"Tn: {round(suma_tn_fila, 2)}\nFOB: {round(suma_fob_fila, 2)}"
        ws.cell(row=fila, column=col).alignment = Alignment(wrap_text=True)

        total_general_toneladas += suma_tn_fila
        total_general_fob += suma_fob_fila

        fila += 1

    # Fila total final para columnas
    ws.cell(row=fila, column=1).value = "Totales"
    ws.cell(row=fila, column=1).font = Font(bold=True)

    col = 2
    for y in years:
        tn_col = total_por_columna[y]['toneladas']
        fob_col = total_por_columna[y]['fob']

        ws.cell(row=fila, column=col).value = round(tn_col, 2)
        ws.cell(row=fila, column=col).font = Font(bold=True)
        ws.cell(row=fila, column=col+1).value = round(fob_col, 2)
        ws.cell(row=fila, column=col+1).font = Font(bold=True)

        col += 2

    # Total general en la última columna
    ws.cell(row=fila, column=col).value = f"Tn: {round(total_general_toneladas, 2)}\nFOB: {round(total_general_fob, 2)}"
    ws.cell(row=fila, column=col).font = Font(bold=True)
    ws.cell(row=fila, column=col).alignment = Alignment(wrap_text=True)

    # Ajustar ancho columnas
    ws.column_dimensions['A'].width = 25
    for i in range(2, 2 + len(years)*2 + 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[col_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={
        'Content-Disposition': 'attachment; filename="reporte_exportaciones.xlsx"',
    })
























def gestion_usuarios(request):
    usuarios = User.objects.all()
    grupos = Group.objects.all()
    mensaje = ''
    tipo_mensaje = ''

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        grupo_nombre = request.POST.get('grupo')

        if username and email and password and grupo_nombre:
            if User.objects.filter(username=username).exists():
                mensaje = 'El nombre de usuario ya existe.'
                tipo_mensaje = 'error'
            else:
                user = User.objects.create_user(username=username, email=email, password=password)
                grupo = Group.objects.get(name=grupo_nombre)
                user.groups.add(grupo)
                mensaje = 'Usuario creado exitosamente.'
                tipo_mensaje = 'success'
        else:
            mensaje = 'Complete todos los campos.'
            tipo_mensaje = 'error'

    return render(request, 'admin.html', {
        'usuarios': usuarios,
        'grupos': grupos,
        'mensaje': mensaje,
        'tipo_mensaje': tipo_mensaje
    })

@csrf_exempt
def editar_usuario(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        username = request.POST.get('username')
        email = request.POST.get('email')
        grupo_nombre = request.POST.get('grupo')

        user = get_object_or_404(User, id=user_id)
        user.username = username
        user.email = email
        user.save()

        if grupo_nombre:
            grupo = Group.objects.get(name=grupo_nombre)
            user.groups.clear()
            user.groups.add(grupo)
        
        return redirect('admin_users')
    
@login_required
def eliminar_usuario(request, pk):
    if request.method == 'POST':
        # Verificamos que el usuario tenga permisos adecuados
        if not request.user.is_superuser:
            return HttpResponseForbidden("No tienes permiso para eliminar usuarios.")

        usuario = get_object_or_404(User, pk=pk)

        # Evitar que el usuario se elimine a sí mismo
        if request.user == usuario:
            return JsonResponse({'error': 'No puedes eliminar tu propio usuario.'}, status=400)

        usuario.delete()
        return JsonResponse({'mensaje': 'Usuario eliminado correctamente.'})

    return HttpResponseBadRequest('Método no permitido.')



def exportaciones_duplicadas_pdf(request):
    # Consulta SQL
    query = """
        SELECT 
            e.id_export      AS id_export_a,
            e."Num_Exped1"   AS num_exped_a,
            e.pedido_comercial_export,
            p.nom_productor_min AS empresa,
            e.fecha_present_export AS fecha_a,
            e2.id_export     AS id_export_b,
            e2."Num_Exped1"  AS num_exped_b,
            e2.fecha_present_export AS fecha_b
        FROM app_exportacion e
        JOIN app_exportacion e2 
            ON e.pedido_comercial_export = e2.pedido_comercial_export
           AND e.id_productor_min_id = e2.id_productor_min_id
           AND e.id_export < e2.id_export
           AND EXTRACT(YEAR FROM e.fecha_present_export) = EXTRACT(YEAR FROM CURRENT_DATE)
           AND EXTRACT(YEAR FROM e2.fecha_present_export) = EXTRACT(YEAR FROM CURRENT_DATE)
           AND e."Estado_anulacion" = false
           AND e2."Estado_anulacion" = false
        JOIN app_prodminero p 
            ON e.id_productor_min_id = p.id_productor_min
        ORDER BY empresa, e.pedido_comercial_export, fecha_a, fecha_b;
    """

    with connection.cursor() as cursor:
        cursor.execute(query)
        rows = cursor.fetchall()

    # Respuesta PDF
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="exportaciones_duplicadas.pdf"'

    # Documento con márgenes ajustados
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        leftMargin=20,
        rightMargin=20,
        topMargin=30,
        bottomMargin=20
    )
    elements = []
    styles = getSampleStyleSheet()

    # Título
    elements.append(Paragraph("Exportaciones con Pedido Comercial Duplicado (Año Actual)", styles["Title"]))
    elements.append(Spacer(1, 12))

    # Encabezados de tabla
    data = [["Cod Export 1", "N° Exped 1", "Fecha 1", "Cod Export 2", "N° Exped 2", "Fecha 2", "Pedido Comercial", "Empresa"]]

    # Agregar filas
    for row in rows:
        data.append([
            row[0],  # id_export_a
            row[1],  # num_exped_a
            row[4].strftime("%d/%m/%Y"),  # fecha_a
            row[5],  # id_export_b
            row[6],  # num_exped_b
            row[7].strftime("%d/%m/%Y"),  # fecha_b
            row[2],  # pedido_comercial_export
            row[3],  # empresa
        ])

    # Definir anchos de columna para que no se pase de margen
    col_widths = [55, 55, 55, 55, 55, 55, 80, 120]

    # Crear tabla
    table = Table(data, repeatRows=1, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),  # texto más chico
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 1), (-1, -1), 4),
        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
    ]))

    elements.append(table)

    # Generar PDF
    doc.build(elements)
    return response



    ################################################CONTRATOS######################################################
def contratos_view(request):
    if request.method == 'POST':
        form = ContratosForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'message': 'Contrato creado correctamente.'})
            messages.success(request, 'Contrato creado correctamente.')
            return redirect('contrato_create')
        else:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': form.errors})
            messages.error(request, 'Por favor corrige los errores en el formulario.')
    else:
        form = ContratosForm()
        concesionarios = Concesionarios.objects.using('catastro').all().order_by('concesionario')
        return render(request, 'contratos/new.html', {'form': form , 'concesionarios':concesionarios})


def lista_contratos(request):
    print('actualizando estados')
    today = timezone.now().date()

    # Contratos actualmente vigentes → activo = True
    Contratos.objects.filter(fecha_ini__lte=today, fecha_fin__gte=today).update(activo=True)

    # Contratos no vigentes (ya finalizaron o aún no comenzaron) → activo = False
    Contratos.objects.exclude(fecha_ini__lte=today, fecha_fin__gte=today).update(activo=False)
    print('actualizados')


    search = request.GET.get('q', '').strip()
    page_number = request.GET.get('page', 1)

    
    subquery = Simsaexpedientescontratos.objects.filter(
        id_contrato=OuterRef('pk')
    ).order_by('id').values('nro_expediente')[:1]

    contratos = Contratos.objects.annotate(
        nro_expediente=Subquery(subquery)
    ).order_by('id')
    
    if search:
        contratos = contratos.filter(
            relacion_id_concesionario__icontains=search
        ) | contratos.filter(
            id__icontains=search
        ) | contratos.filter(
            expedientes__nro_expediente__icontains=search
        ).distinct()

    paginator = Paginator(contratos, 10)  # 10 contratos por página
    page_obj = paginator.get_page(page_number)

    # Si es una solicitud AJAX, devolver JSON
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        data = []
        for contrato in page_obj:
            expediente = contrato.expedientes.first()  
            data.append({
                'id': contrato.id,
                'nro_expediente': expediente.nro_expediente if expediente else '',
                'concesionario': contrato.relacion_id_concesionario,
                'pago_cano': contrato.pago_cano,
                'pago_regalias': contrato.pago_regalias,
                'opcion_compra': contrato.opcion_compra,
                'activo': contrato.activo,
            })

        return JsonResponse({
            'contratos': data,
            'num_pages': paginator.num_pages,
            'current_page': page_obj.number,
        })

    return render(request, 'contratos/list.html', {
        'contratos': page_obj,
        'search': search,
    })


def str_to_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ['sí', 'si', 'true', '1']
    return False

def crear_contrato(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Manejo de booleano 'activo'
            activo = str(data.get('activo', '')).lower() == 'si'

            contrato = Contratos.objects.using('catastro').create(
                expediente=data.get('nro_expediente'),
                id_concesionario=data.get('relacion_id_concesionario'),
                paga_canon=data.get('pago_cano', False),
                opcion_compra=data.get('opcion_compra', False),
                mineral_explotacion=data.get('mineral_explotacion'),
                activo=activo,
                fecha_ini=data.get('fecha_ini') or None,
                fecha_fin=data.get('fecha_fin') or None,
                createby=data.get('createby', 'Cargador'),  # se setea "a mano"
                createdate=data.get('createdate') or timezone.now().date()  # se setea "a mano"
            )

            return JsonResponse({'message': 'Contrato creado correctamente', 'id': contrato.id}, status=201)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Método no permitido'}, status=405)


def consultar_contratos_por_expediente(request):
    expediente = request.GET.get('expediente', '').strip()
    contratos = []
    cantera_cateo_minas = []
    grupos_mineros = []

    if expediente:
        concesionario_subquery = Concesionarios.objects.using('catastro').filter(
            concesionario=OuterRef('relacion_id_concesionario')
        ).values('concesionario')[:1]
        
        contratos = Contratos.objects.filter(
            id__in=Simsaexpedientescontratos.objects.using('catastro').filter(
                nro_expediente__icontains=expediente
            ).values_list('contrato_id', flat=True)
        ).annotate(
            nombre_concesionario=Subquery(concesionario_subquery)
        )

        cantera_cateo_minas = CanteraCateoMina.objects.using('catastro').filter(expediente__icontains=expediente)
        for item in cantera_cateo_minas:
            item.concesionario = item.concesionario.replace('"','')
            item.concesionario = item.concesionario.replace('{','')
            item.concesionario = item.concesionario.replace('}','')
            item.mineral = item.mineral.replace('"','')
            item.mineral = item.mineral.replace('{','')
            item.mineral = item.mineral.replace('}','')

        grupos_mineros = GrupoMinero.objects.using('catastro').filter(expediente__icontains=expediente)

    context = {
        'serch': True if expediente else False,
        'expediente': expediente,
        'contratos': contratos,
        'cantera_cateo_minas': cantera_cateo_minas,
        'grupos_mineros': grupos_mineros,
    }
    return render(request, 'contratos/serch.html', context)


def consulta_expediente_view(request):
    expediente = request.GET.get('expediente', '').strip()
    contratos = []
    cantera_cateo_minas = []
    grupos_mineros = []

    if expediente:
        concesionario_subquery = Concesionarios.objects.using('catastro').filter(
            concesionario=OuterRef('relacion_id_concesionario')
        ).values('id')[:1]
        
        contratos = Contratos.objects.filter(
            id__in=Simsaexpedientescontratos.objects.filter(
                nro_expediente__icontains=expediente
            ).values_list('contrato_id', flat=True)
        ).annotate(
            nombre_concesionario=Subquery(concesionario_subquery)
        )
        
        cantera_cateo_minas = CanteraCateoMina.objects.using('catastro').filter(expediente__icontains=expediente)
        grupos_mineros = GrupoMinero.objects.using('catastro').filter(expediente__icontains=expediente)
        concesionarios = Concesionarios.objects.using('catastro').all().order_by('concesionario')
        print('hola')
        for item in concesionarios:
            print('concesionario',item.concesionario)
    context = {
        'expediente': expediente,
        'contratos': contratos,
        'cantera_cateo_minas': cantera_cateo_minas,
        'grupos_mineros': grupos_mineros,
    }
    return render(request, 'contratos/consulta_resultados.html', context)



def verificar_expediente(request, nro):
    expediente = None

    try:
        expediente = CanteraCateoMina.objects.using('catastro').get(expediente=nro)
        estado = (expediente.estado or "").strip().lower()
        nombre = expediente.nombre
        if expediente.tipo == 'minas':
            tipo = 'Mina'
        elif expediente.tipo == 'canteras':
            tipo = 'Cantera'
        else:
            tipo = expediente.tipo
    except CanteraCateoMina.DoesNotExist:
        try:
            expediente = GrupoMinero.objects.using('catastro').get(expediente=nro)
            estado = (expediente.estado or "").strip().lower()
            nombre = expediente.nombre
            tipo = 'Grupo Minero'

        except GrupoMinero.DoesNotExist:
            return HttpResponse(status=404)  # No encontrado en ninguna de las dos

    

    if estado == "vigente":
        return JsonResponse({"nombre": nombre, "tipo": tipo}, status=200)
    elif estado == "archivo":
        return JsonResponse({"nombre": nombre, "tipo": tipo}, status=300)
    else:
        return HttpResponse("Estado desconocido", status=500)


########################################ADMINISTRACION################################
def admin_home(request):

    query = request.GET.get('q', '')
    usuarios = User.objects.all()
    organismos = Organismo.objects.all()

    if query:
        usuarios = usuarios.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )

    usuarios = usuarios.order_by('username')
    return render(request, 'administracion/usuarios.html', {
        'usuarios': usuarios,
        'query': query,
        'organismos':organismos
    })

def check_username(request):
    username = request.GET.get('username', '').strip()
    exists = User.objects.filter(username=username).exists()
    return JsonResponse({'exists': exists})

def crear_usuario(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        print(data)
        username = data.get('username', '').strip()
        password = data.get('usuariopass', '')
        nombre = data.get('usuarionom', '').strip()
        email = data.get('usuarioemail', '').strip()
        celular = data.get('usuariocelular', '').strip()
        organismos_ids = data.get('organismos', [])

        # Validaciones
        if not username or ' ' in username or User.objects.filter(username=username).exists():
            return JsonResponse({'success': False, 'error': 'Usuario inválido o ya existe'})

        # Crear usuario
        user = User.objects.create_user(username=username, password=password, email=email, first_name=nombre)
        user.save()  # Esto disparará tu signal para crear PerfilUsuario automáticamente

        # Asignar organismos
        organismos_asignados = []
        for org_id in organismos_ids:
            try:
                org = Organismo.objects.get(organismoid=org_id)
                OrganismoUsuario.objects.create(usuario=user, organismo=org)
                organismos_asignados.append(org.organismonombre)
            except Organismo.DoesNotExist:
                continue  # Ignorar si no existe el organismo
        try:
            organismo_str = ", ".join(organismos_asignados) if organismos_asignados else "Ninguno asignado"

            subject = f"Bienvenido/a al {getattr(settings, 'ORGANISMO', 'Secretaría de Minería y Energía')}"
            message = (
                f"Estimado/a {nombre},\n\n"
                f"Le damos la bienvenida a la {getattr(settings, 'ORGANISMO', 'Secretaría de Minería y Energía')}.\n\n"
                f"Puede acceder al sistema en la siguiente URL:\n"
                f"{getattr(settings, 'SITE_URL', 'http://localhost:8000')}\n\n"
                f"Sus credenciales de acceso son:\n"
                f"Usuario: {username}\n"
                f"Contraseña Temporal: {password} \n\n"
                f"Organismo(s) asignado(s): {organismo_str}\n\n"
                f"Por favor, conserve estos datos de manera segura.\n\n"
                f"Atentamente,\n"
                f"{getattr(settings, 'ORGANISMO', 'Secretaría de Minería y Energía')}"
            )

            send_mail(
                subject,
                message,
                getattr(settings, 'DEFAULT_FROM_EMAIL', 'no-reply@mineria.gob.ar'),
                [email],
                fail_silently=False,
            )
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Usuario creado, pero no se pudo enviar el email: {str(e)}'})

        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Método no permitido'})

def get_usuario_data(request, user_id):
    usuario = get_object_or_404(User, id=user_id)

    # organismos asociados a este usuario
    organismos_usuario_ids = list(usuario.organismos_usuario.values_list("organismo_id", flat=True))

    # todos los organismos (opcional para dropdown)
    todos_organismos = [
        {
            "id": o.organismoid,
            "nombre": o.organismonombre,
            "asociado": o.organismoid in organismos_usuario_ids
        }
        for o in Organismo.objects.all()
    ]

    # perfil
    perfil = getattr(usuario, 'perfilusuario', None)

    data = {
        "success": True,
        "user": {
            "id": usuario.id,
            "username": usuario.username,
            "nombre": usuario.first_name + ' ' + usuario.last_name if usuario else '',
            "email": usuario.email if perfil else usuario.email,
            "celular": perfil.usuariocelular if perfil else "",
            "organismos": organismos_usuario_ids
        },
        "todos_organismos": todos_organismos
    }

    return JsonResponse(data)

@login_required
def toggle_usuario_activo(request, user_id):
    usuario = get_object_or_404(User, pk=user_id)
    usuario.is_active = not usuario.is_active
    usuario.save()
    estado = "activado" if usuario.is_active else "desactivado"
    messages.success(request, f"✅ Usuario {usuario.username} {estado}.")
    return redirect('administration')

@login_required
@require_POST
def editar_usuario_modal(request, user_id):
    usuario = get_object_or_404(User, pk=user_id)
    usuario.first_name = request.POST.get('first_name', '')
    usuario.last_name = request.POST.get('last_name', '')
    usuario.email = request.POST.get('email', '')
    usuario.save()

    messages.success(request, f"✏️ Usuario {usuario.username} actualizado correctamente.")
    return redirect('administration')

def editar_usuario_view(request, user_id):
    usuario = get_object_or_404(User, pk=user_id)
    perfil, _ = PerfilUsuario.objects.get_or_create(user=usuario)

    if request.method == "POST":
        try:
            
            # Actualizar datos User
            usuario.first_name = request.POST.get('usuarionom', usuario.first_name)
            usuario.last_name = request.POST.get('last_name', usuario.last_name)
            usuario.email = request.POST.get('usuarioemail', usuario.email)
            usuario.save()

            # Actualizar datos PerfilUsuario
            perfil.usuariopin = request.POST.get('usuariopin', perfil.usuariopin)
            perfil.usuarioemail = request.POST.get('usuarioemail', perfil.usuarioemail)
            perfil.usuarioestado = request.POST.get('usuarioestado', perfil.usuarioestado)
            perfil.usuariocelular = request.POST.get('usuariocelular', perfil.usuariocelular)
            perfil.usuarionom = request.POST.get('usuarionom', perfil.usuarionom)
            perfil.save()

            # Actualizar organismos
            organismos_seleccionados = request.POST.getlist('organismos')

            # Eliminamos todos los que no están seleccionados
            OrganismoUsuario.objects.filter(usuario=usuario).exclude(
                organismo_id__in=organismos_seleccionados
            ).delete()

            # Añadimos los nuevos seleccionados (si no existen)
            for org_id in organismos_seleccionados:
                org = Organismo.objects.filter(pk=org_id).first()
                if org:
                    OrganismoUsuario.objects.get_or_create(usuario=usuario, organismo=org)

            return JsonResponse({
                "success": True,
                "message": f"✏️ Usuario {usuario.username} actualizado correctamente."
            })

        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e)
            }, status=400)

    return JsonResponse({
        "success": False,
        "error": "Método no permitido"
    }, status=405)
@login_required
def blanquear_contraseña(request, user_id):
    usuario = get_object_or_404(User, pk=user_id)
    nueva_password = get_random_string(length=8)
    usuario.set_password(nueva_password)
    usuario.save()
    send_mail(
                subject='🔐 Recuperación de contraseña – Sistema de la Secretaria de Mineria',
                message=(
                    f"Estimado/a {usuario.username},\n\n"
                    f"Desde la Secretaría de Minería y Energía de la Provincia de Salta le informamos que se ha generado una nueva contraseña temporal para su acceso al Sistema.\n\n"
                    f"🔑 Nueva contraseña temporal: {nueva_password}\n\n"
                    f"Por razones de seguridad, le recomendamos ingresar al sistema lo antes posible y modificar su contraseña desde su perfil.\n\n"
                    f"Si usted no solicitó esta recuperación o tiene alguna dificultad para acceder, por favor comuníquese con nuestra Mesa de Ayuda escribiendo a: secretariademineriadesalta@gmail.com\n\n"
                    f"Atentamente,\n"
                    f"Secretaría de Minería y Energía\n"
                    f"Gobierno de la Provincia de Salta"
                ),
                from_email=None,  # Usa DEFAULT_FROM_EMAIL si está configurado en settings.py
                recipient_list=[usuario.email],
                fail_silently=False,
            )

    try:
        perfil = usuario.perfilusuario
        perfil.debe_cambiar_password = True
        perfil.save()
    except PerfilUsuario.DoesNotExist:
        pass  # Si no tiene perfil, no pasa nada

    messages.success(request, f"🔐 Contraseña de {usuario.username} blanqueada. Se pedirá cambio en el próximo acceso.")
    return redirect('administration')

@login_required
def editar_usuario(request, user_id):
    # Implementá esto si querés un formulario de edición
    messages.info(request, f"🛠️ Funcionalidad de edición aún no implementada.")
    return redirect('administration')

def generar_password_seguro(longitud=10):
    chars = string.ascii_letters + string.digits + "!@#$%^&*"
    return ''.join(secrets.choice(chars) for _ in range(longitud))

def crear_usuario_view(request):
    organismos = Organismo.objects.all()
    context = {
        'organismos':organismos
    }
    return render(request, 'administracion/crear_usuario.html',context)








#####################################################PROVEEDORES#######################################################

RAZONES_SOCIALES_REQUERIDAS = [
    "ANDINA PERFORACIONES SRL", "CATERING DE ALTURA S.A.S", "EFE-BUS SRL", "FERRIL CARLOS JAVIER",
    "HG COMUNICACIONES SRL", "LUMI PUNA SRL", "MORON ARANSAY OSCAR ALFREDO", "MVA SRL",
    "NOA GENERACION SRL", "PLANETA PUNA SRL", "SERMINCA SRL", "SERVICIOS Y EXPLOTACIONES MINERAS CRUZ",
    "SMART WASTE SAS", "SOSA GUEVARA NESTOR ABEL", "HIDROTEC SRL", "JSC MINNING S.A.S", "BIOTEC SRL",
    "MANUFACTURA DE LOS ANDES SA", "SARAPURA WALTER SERGIO", "Salud Integrada S.R.L.", "Lumi People SRL",
    "Pacha Consultora Ambiental SRL", "Arqueoambiental SA", "HYUNDAI HEAVY INDUSTRIES ARGENTINA SRL",
    "TEPSI S.A.", "ANDDES ARGENTINA S.A", "BSD Ingeniería y Servicios SA", "AGGREKO ARGENTINA S.R.L.",
    "Posco Eco & Challenge CO. LTD. S.E.E.", "MINERA MARIMARI S.A.", "SANATORIO PASTEUR SA",
    "SAMJIN-FORTIS UTE", "MILICIC -AGV UTE", "BMI S.A.", "Todo obras SRL - Samjin Electric SA U.T.",
    "BBC UTE", "GEOTEC SRL"
]

# Normalizador
@login_required
def normalizar(texto):
    texto = texto.upper().strip()
    texto = unicodedata.normalize('NFD', texto)
    texto = ''.join(c for c in texto if unicodedata.category(c) != 'Mn')  # eliminar tildes
    texto = texto.replace('.', '').replace('-', ' ')
    texto = ' '.join(texto.split())
    return texto

@login_required
def verificar_proveedores(request):
    
    # Validar y obtener datos del formulario
    anio_param = request.POST.get("anio")
    empresa = request.POST.get("empresa", "").strip()
    inspector = request.POST.get("controlador", "").strip()
    codigo = request.POST.get("codigo_inspeccion", "").strip()
    columna = request.POST.get("columna_proveedores", "").strip()
    observaciones = request.POST.get("observaciones", "").strip()
    archivo_excel = request.FILES.get("archivo_excel")

    # Validación estricta
    errores = []
    if not empresa:
        errores.append("Debe ingresar el nombre de la empresa.")
    if not inspector:
        errores.append("Debe ingresar el nombre del inspector.")
    if not columna or not columna.isdigit():
        errores.append("Debe ingresar un número de columna válido.")
    if not archivo_excel:
        errores.append("Debe subir un archivo Excel válido.")

    if errores:
        return JsonResponse({"error": "Faltan datos obligatorios.", "detalle": errores}, status=400)

    columna_proveedores = int(columna)

    try:
        columna_proveedores = int(columna)
    except ValueError:
        return JsonResponse({"error": "La columna de proveedores debe ser un número entero."}, status=400)

    # Registrar inspección
    inspeccion = InspeccionProveedores.objects.create(
        empresa=empresa,
        anio=anio_param or None,
        inspector=inspector,
        codigo_inspeccion=codigo or None,
        columna_proveedores=columna_proveedores,
        archivo_excel=archivo_excel,
        observaciones=observaciones or None,
        usuario_registro=request.user
    )

    # Filtro de base legal
    registros = RegistroProveedores.objects.all().order_by(id)
    if anio_param:
        try:
            anio = int(anio_param)
            registros = registros.filter(anio=anio)
        except ValueError:
            return JsonResponse({"error": "El parámetro 'anio' debe ser un número entero."}, status=400)

    # Armar base normalizada
    base = {
        normalizar(r.nombre_razon_social): {
            "id": r.id,
            "nombre_db": r.nombre_razon_social
        }
        for r in registros
    }

    # Comparar con los nombres requeridos (de tu lógica previa)
    nombres_referencia = [normalizar(n) for n in RAZONES_SOCIALES_REQUERIDAS]
    encontrados = []
    no_encontrados = []

    for nombre in nombres_referencia:
        coincidencias = get_close_matches(nombre, base.keys(), n=1, cutoff=0.85)
        if coincidencias:
            match = coincidencias[0]
            encontrados.append({
                "referencia": nombre,
                "coincidencia": base[match]["nombre_db"],
                "id": base[match]["id"]
            })
        else:
            no_encontrados.append(nombre)

    total = len(nombres_referencia)
    porcentaje = round(len(encontrados) * 100 / total, 2) if total > 0 else 0

    return JsonResponse({
        "mensaje": "Inspección registrada y comparación realizada con éxito.",
        "inspeccion": {
            "id": inspeccion.id,
            "empresa": inspeccion.empresa,
            "inspector": inspeccion.inspector,
            "codigo": inspeccion.codigo_inspeccion,
            "columna": inspeccion.columna_proveedores,
            "usuario": request.user.username,
            "fecha": inspeccion.fecha_registro.strftime("%d/%m/%Y %H:%M")
        },
        "anio": anio_param or "TODOS",
        "total_referencia": total,
        "coincidencias": len(encontrados),
        "faltantes": len(no_encontrados),
        "porcentaje_encontrados": f"{porcentaje}%",
        "encontrados": encontrados,
        "no_encontrados": no_encontrados
    }, json_dumps_params={"ensure_ascii": False, "indent": 2})

def normalizar_nombres(texto):
    texto = texto.upper().strip()
    texto = unidecode(texto)
    texto = texto.replace('.', '').replace('-', ' ')
    return ' '.join(texto.split())

@login_required
def proveedores_list(request):
    registros = RegistroProveedores.objects.all()

    # Filtros GET
    anio = request.GET.get("anio")
    estado = request.GET.get("estado")
    tipo = request.GET.get("tipo_registro")
    nombre = request.GET.get("nombre")
    estado_segun_fecha = request.GET.get("estado_fecha")

    print(estado_segun_fecha)

    hoy = datetime.now()

    if estado_segun_fecha == "1":
        registros = registros.filter(fecha_alta__lte=hoy, fecha_vto__gte=hoy)
    elif estado_segun_fecha == "0":
        registros = registros.exclude(fecha_alta__lte=hoy, fecha_vto__gte=hoy)
    if anio:
        registros = registros.filter(anio=anio)
    if estado:
        registros = registros.filter(estado__icontains=estado)
    if tipo:
        registros = registros.filter(tipo_registro__icontains=tipo)
    if nombre:
        registros = registros.filter(nombre_razon_social__icontains=nombre)

    # Para los select
    anios_disponibles = RegistroProveedores.objects.order_by().values_list('anio', flat=True).distinct()
    tipos_disponibles = RegistroProveedores.objects.order_by().values_list('tipo_registro', flat=True).distinct()
    estados_disponibles = RegistroProveedores.objects.order_by().values_list('estado', flat=True).distinct()

    return render(request, "proveedores/list.html", {
        "registros": registros,
        "anios": sorted(filter(None, set(anios_disponibles))),
        "tipos": sorted(filter(None, set(tipos_disponibles))),
        "estados": sorted(filter(None, set(estados_disponibles))),
        "estado_fecha":estado_segun_fecha
    })

def actualizar_proveedores(request):
    if request.method == "POST" and request.FILES.get("excel_file"):
        excel_file = request.FILES["excel_file"]

        try:
            # Leer Excel
            df = pd.read_excel(excel_file)

            # Funciones auxiliares
            def parse_fecha(val):
                if pd.isna(val):
                    return None
                if isinstance(val, datetime):
                    return val.date()
                try:
                    return pd.to_datetime(val).date()
                except Exception:
                    return None

            def parse_datetime(val):
                if pd.isna(val):
                    return None
                if isinstance(val, datetime):
                    return val
                try:
                    return pd.to_datetime(val)
                except Exception:
                    return None

            def parse_bool(val):
                if pd.isna(val):
                    return None
                v = str(val).strip().lower()
                if v in ["sí", "si", "x", "true", "1"]:
                    return True
                if v in ["no", "false", "0"]:
                    return False
                return None

            with transaction.atomic():
                # 🔥 Borrar datos actuales
                RegistroProveedores.objects.all().delete()

                proveedores = []
                for _, row in df.iterrows():
                    proveedor = RegistroProveedores(
                        creado=parse_datetime(row.get("Creado")),
                        numero_tramite=row.get("N° TRÁMITE"),
                        mes=row.get("Mes"),
                        anio=row.get("Año") if not pd.isna(row.get("Año")) else None,
                        numero_certificado=row.get("N° de Certificado"),
                        tipo_registro=row.get("TIPO DE REGISTRO"),
                        tramite=row.get("TRÁMITE"),
                        fecha_alta=parse_fecha(row.get("FECHA ALTA")),
                        fecha_vto=parse_fecha(row.get("Fecha de VTO")),
                        estado=row.get("ESTADO"),
                        numero_expediente=row.get("N° de expediente"),
                        nombre_razon_social=row.get("APELLIDO Y NOMBRE O RAZÓN SOCIAL"),
                        cuit_cuil=row.get("CUIT – CUIL"),
                        domicilio_real=row.get("DOMICILIO REAL"),
                        domicilio_social=row.get("DOMICILIO SOCIAL"),
                        localidad=row.get("LOCALIDAD"),
                        domicilio_fiscal=row.get("DOMICILIO FISCAL"),
                        telefono=row.get("TELÉFONO"),
                        representante_legal=row.get("DATOS DEL REPRESENTANTE LEGAL"),
                        documento_identidad=row.get("DOCUMENTO DE IDENTIDAD"),
                        correo_electronico=row.get("CORREO ELECTRÓNICO") or None,
                        actividad=row.get("ACTIVIDAD"),
                        camara=row.get("CAMARA QUE INTEGRA"),
                        declaracion_jurada=parse_bool(row.get("DECLARACIÓN JURADA")),
                        persona_asignada=row.get("PERSONA ASIGNADA"),
                        nomina_trabajadores=parse_bool(row.get("¿NOMINA TRABAJADORES?")),
                        fotocopia_dni=parse_bool(row.get("¿FOTOCOPIA DNI?")),
                        certificado_residencia=parse_bool(row.get("¿CERTIFICADO DE RESIDENCIA?")),
                        inscripcion_afip_dgr=parse_bool(row.get("¿INSCRIPCION AFIP / DGR?")),
                        regularizacion_fiscal=parse_bool(row.get("¿REGULARIZACIÓN FISCAL (F500)?")),
                        contrato_social=parse_bool(row.get("¿CONTRATO SOCIAL?")),
                        acta_designacion_autoridades=parse_bool(row.get("¿ACTA DESIGNACIÓN AUTORIDADES?")),
                        dni_autoridades=parse_bool(row.get("¿DNI DE AUTORIDADES?")),
                        certificado_residencia_autoridades=parse_bool(row.get("¿CERTIFICADO RESIDENCIA DE AUTORIDADES?")),
                        dni_representante=parse_bool(row.get("¿DNI REPRESENTANTE?")),
                        certificado_residencia_representante=parse_bool(row.get("¿CERTIFICADO DE RESIDENCIA REPRESENTANTE?")),
                        poder_otorgado_representante=parse_bool(row.get("¿PODER OTORGADO AL REPRESENTANTE?")),
                        f931_seguridad_social=parse_bool(row.get("¿F931 - SEGURIDAD SOCIAL?")),
                        constancia_matriculacion_trabajadores=parse_bool(row.get("¿CONSTANCIA DE MATRICULACIÓN TRABAJADORES?")),
                        constancia_cuil=parse_bool(row.get("¿CONSTANCIA DE CUIL?")),
                    )
                    proveedores.append(proveedor)

                RegistroProveedores.objects.bulk_create(proveedores)

            messages.success(request, f"✅ Se cargaron {len(proveedores)} proveedores correctamente")

        except Exception as e:
            messages.error(request, f"❌ Error al actualizar proveedores: {str(e)}")
    return render(request,'proveedores/actualizar_proveedores.html')

@login_required
def detalle_proveedor(request, pk):
    proveedor = get_object_or_404(RegistroProveedores, pk=pk)
    documentos_presentados = [
        ('Declaración Jurada', proveedor.declaracion_jurada),
        ('Nómina de Trabajadores', proveedor.nomina_trabajadores),
        ('Fotocopia DNI', proveedor.fotocopia_dni),
        ('Cert. Residencia', proveedor.certificado_residencia),
        ('Inscripción AFIP/DGR', proveedor.inscripcion_afip_dgr),
        ('Regularización Fiscal', proveedor.regularizacion_fiscal),
        ('Contrato Social', proveedor.contrato_social),
        ('Acta Designación Autoridades', proveedor.acta_designacion_autoridades),
        ('DNI Autoridades', proveedor.dni_autoridades),
        ('Cert. Residencia Autoridades', proveedor.certificado_residencia_autoridades),
        ('DNI Representante', proveedor.dni_representante),
        ('Cert. Residencia Representante', proveedor.certificado_residencia_representante),
        ('Poder Representante', proveedor.poder_otorgado_representante),
        ('F931 Seguridad Social', proveedor.f931_seguridad_social),
        ('Constancia Matriculación', proveedor.constancia_matriculacion_trabajadores),
        ('Constancia CUIL', proveedor.constancia_cuil),
    ]

    return render(request, 'proveedores/detalle.html', {
        'proveedor': proveedor,
        'documentos_presentados': documentos_presentados
    })

@login_required
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(RegistroProveedores, pk=pk)

    if request.method == "POST":
        # Text/char/date inputs
        proveedor.creado = request.POST.get("creado") or proveedor.creado
        proveedor.numero_tramite = request.POST.get("numero_tramite")
        proveedor.mes = request.POST.get("mes")
        proveedor.anio = request.POST.get("anio") or None
        proveedor.numero_certificado = request.POST.get("numero_certificado")
        proveedor.tipo_registro = request.POST.get("tipo_registro")
        proveedor.tramite = request.POST.get("tramite")
        proveedor.fecha_alta = request.POST.get("fecha_alta") or None
        proveedor.fecha_vto = request.POST.get("fecha_vto") or None
        proveedor.estado = request.POST.get("estado")
        proveedor.numero_expediente = request.POST.get("numero_expediente")
        proveedor.nombre_razon_social = request.POST.get("nombre_razon_social")
        proveedor.cuit_cuil = request.POST.get("cuit_cuil")
        proveedor.domicilio_real = request.POST.get("domicilio_real")
        proveedor.domicilio_social = request.POST.get("domicilio_social")
        proveedor.localidad = request.POST.get("localidad")
        proveedor.domicilio_fiscal = request.POST.get("domicilio_fiscal")
        proveedor.telefono = request.POST.get("telefono")
        proveedor.representante_legal = request.POST.get("representante_legal")
        proveedor.documento_identidad = request.POST.get("documento_identidad")
        proveedor.correo_electronico = request.POST.get("correo_electronico")
        proveedor.actividad = request.POST.get("actividad")
        proveedor.camara = request.POST.get("camara")
        proveedor.persona_asignada = request.POST.get("persona_asignada")

        # Boolean fields (checked if present)
        proveedor.declaracion_jurada = 'declaracion_jurada' in request.POST
        proveedor.nomina_trabajadores = 'nomina_trabajadores' in request.POST
        proveedor.fotocopia_dni = 'fotocopia_dni' in request.POST
        proveedor.certificado_residencia = 'certificado_residencia' in request.POST
        proveedor.inscripcion_afip_dgr = 'inscripcion_afip_dgr' in request.POST
        proveedor.regularizacion_fiscal = 'regularizacion_fiscal' in request.POST
        proveedor.contrato_social = 'contrato_social' in request.POST
        proveedor.acta_designacion_autoridades = 'acta_designacion_autoridades' in request.POST
        proveedor.dni_autoridades = 'dni_autoridades' in request.POST
        proveedor.certificado_residencia_autoridades = 'certificado_residencia_autoridades' in request.POST
        proveedor.dni_representante = 'dni_representante' in request.POST
        proveedor.certificado_residencia_representante = 'certificado_residencia_representante' in request.POST
        proveedor.poder_otorgado_representante = 'poder_otorgado_representante' in request.POST
        proveedor.f931_seguridad_social = 'f931_seguridad_social' in request.POST
        proveedor.constancia_matriculacion_trabajadores = 'constancia_matriculacion_trabajadores' in request.POST
        proveedor.constancia_cuil = 'constancia_cuil' in request.POST

        proveedor.save()

        # Si la petición vino por fetch (AJAX), devolver HTML parcial del detalle
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            html = render_to_string("proveedores/detalle.html", {"proveedor": proveedor}, request=request)
            return HttpResponse(html)

        return redirect("detalle_proveedor", pk=proveedor.pk)

    return render(request, "proveedores/detalle.html", {"proveedor": proveedor})
# Función para normalizar y tokenizar nombres
SUFIJOS_LEGALES = ["srl", "sa", "sas", "spa", "sc", "s.a.", "s.a.s."]
def tokenize_nombre(nombre):
    nombre = str(nombre).lower().strip()
    nombre = ''.join(
        c for c in unicodedata.normalize('NFKD', nombre)
        if not unicodedata.combining(c) and (c.isalnum() or c.isspace())
    )
    tokens = nombre.split()
    # Eliminar sufijos legales del final
    while tokens and tokens[-1] in SUFIJOS_LEGALES:
        tokens.pop()
    return tokens

def nombres_coinciden(nombre_excel, nombre_base, cutoff=0.8):
    tokens_excel = tokenize_nombre(nombre_excel)
    tokens_base = tokenize_nombre(nombre_base)

    coincidencias = 0
    for te in tokens_excel:
        for tb in tokens_base:
            if te == tb or (len(te) == 1 and tb.startswith(te)) or (len(tb) == 1 and te.startswith(tb)):
                coincidencias += 1
                break

    total = max(len(tokens_excel), len(tokens_base))
    similitud = coincidencias / total if total else 0
    return similitud >= cutoff


def generar_codigo_unico():
    while True:
        random_str = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
        codigo = f"INS-{datetime.now().strftime('%Y%m%d')}-{random_str}"
        if not InspeccionProveedores.objects.filter(codigo_inspeccion=codigo).exists():
            return codigo

@login_required
def comparar_proveedores_excel(request):
    '''if request.method != "POST":
        return JsonResponse({"error": "Método no permitido. Debe ser POST."}, status=405)

    archivo = request.FILES.get("archivo_excel")
    fecha_inspeccion = request.POST.get("fecha_inspeccion")
    empresa = request.POST.get("empresa", "").strip()
    inspector = request.POST.get("controlador", "").strip()
    columna_proveedores = request.POST.get("columna_proveedores", "").strip()
    columna_dni = request.POST.get("columna_dni", "").strip()
    observaciones = request.POST.get("observaciones", "").strip()

    errores = []

    if fecha_inspeccion:
        try:
            fecha_obj = datetime.strptime(fecha_inspeccion, "%Y-%m-%d").date()
        except ValueError:
            errores.append("Formato de fecha_inspeccion inválido, debe ser YYYY-MM-DD.")
            fecha_obj = None
    else:
        fecha_obj = None

    if not archivo:
        errores.append("Debe adjuntar un archivo Excel.")
    if not empresa:
        errores.append("Debe ingresar el nombre de la empresa.")
    if not inspector:
        errores.append("Debe ingresar el nombre del inspector.")
    if not columna_proveedores.isdigit():
        errores.append("Debe ingresar un número de columna válido para proveedores.")
    if not columna_dni.isdigit():
        errores.append("Debe ingresar un número de columna válido para DNI.")

    if errores:
        return JsonResponse({"error": "Faltan datos obligatorios.", "detalle": errores}, status=400)

    columna_proveedores = int(columna_proveedores)
    columna_dni = int(columna_dni)

    try:
        hojas = pd.read_excel(archivo, engine='openpyxl', sheet_name=None, skiprows=1)
    except Exception as e:
        return JsonResponse({"error": f"Error al leer el Excel: {str(e)}"}, status=400)

    registros = RegistroProveedores.objects.exclude(
        nombre_razon_social__isnull=True
    ).exclude(
        nombre_razon_social__exact=""
    )

    vigentes, vencidos, no_encontrados = [], [], []
    vigentes_set, vencidos_set, no_encontrados_set = set(), set(), set()
    empleados_por_proveedor = {}
    proveedor_dni_set = {}
    total_set = set()

    for nombre_hoja, df in hojas.items():
        if columna_proveedores >= len(df.columns) or columna_dni >= len(df.columns):
            continue

        mask_valid = df.iloc[:, columna_proveedores].notna() & df.iloc[:, columna_dni].notna()
        proveedores_raw = df.loc[mask_valid, df.columns[columna_proveedores]].astype(str).tolist()
        dnis_raw = df.loc[mask_valid, df.columns[columna_dni]].astype(str).tolist()

        coincidencias_confirmadas = []
        sin_coincidencia = []

        for i, proveedor_raw in enumerate(proveedores_raw):
            nombre_excel = proveedor_raw.strip()
            if "empresa" in nombre_excel.lower():
                continue

            total_set.add(nombre_excel)

            # Buscar todos los registros que coincidan
            registros_match = [r for r in registros if nombres_coinciden(nombre_excel, r.nombre_razon_social, cutoff=0.8)]

            if registros_match:
                # Tomar el registro más reciente según fecha_vto
                proveedor = max(registros_match, key=lambda x: x.fecha_vto or datetime.min.date())
                proveedor_nombre = proveedor.nombre_razon_social

                if proveedor_nombre not in vigentes_set and proveedor_nombre not in vencidos_set:
                    empleados = proveedor_dni_set.get(proveedor_nombre, set())
                    entry = {
                        "referencia": proveedores_raw[i],
                        "coincidencia": proveedor_nombre,
                        "id": proveedor.id,
                        "fecha_alta": proveedor.fecha_alta.strftime("%d/%m/%Y") if proveedor.fecha_alta else None,
                        "fecha_vto": proveedor.fecha_vto.strftime("%d/%m/%Y") if proveedor.fecha_vto else None,
                        "hoja": nombre_hoja,
                        "empleados": len(empleados)
                    }

                    if fecha_obj and proveedor.fecha_alta and proveedor.fecha_vto and proveedor.fecha_alta <= fecha_obj <= proveedor.fecha_vto:
                        vigentes.append(entry)
                        vigentes_set.add(proveedor_nombre)
                    else:
                        vencidos.append(entry)
                        vencidos_set.add(proveedor_nombre)

                coincidencias_confirmadas.append((proveedor_nombre, dnis_raw[i]))
            else:
                sin_coincidencia.append((nombre_excel, dnis_raw[i]))

        # Agrupar no encontrados por similitud
        grupos = {}
        for nombre, dni in sin_coincidencia:
            clave = get_close_matches(nombre, grupos.keys(), n=1, cutoff=0.8)
            if clave:
                grupos[clave[0]].add(dni)
            else:
                grupos[nombre] = set([dni])

        for grupo, dnis in grupos.items():
            empleados_por_proveedor[grupo] = len(dnis)
            if grupo not in no_encontrados_set:
                no_encontrados.append({
                    "referencia": grupo,
                    "hoja": "-",
                    "empleados": len(dnis)
                })
                no_encontrados_set.add(grupo)

        for proveedor_match, dni in coincidencias_confirmadas:
            if proveedor_match not in empleados_por_proveedor:
                empleados_por_proveedor[proveedor_match] = 0
                proveedor_dni_set[proveedor_match] = set()
            if dni not in proveedor_dni_set[proveedor_match]:
                proveedor_dni_set[proveedor_match].add(dni)
                empleados_por_proveedor[proveedor_match] += 1

    total = len(total_set)

    for item in vigentes:
        item["empleados"] = empleados_por_proveedor.get(item["coincidencia"], 0)
    for item in vencidos:
        item["empleados"] = empleados_por_proveedor.get(item["coincidencia"], 0)

    inspeccion = InspeccionProveedores.objects.create(
        empresa=empresa,
        fecha_inspeccion=fecha_inspeccion or None,
        anio=fecha_obj.year if fecha_obj else None,
        inspector=inspector,
        codigo_inspeccion=generar_codigo_unico(),
        columna_proveedores=columna_proveedores,
        columna_dni=columna_dni,
        archivo_excel=archivo,
        observaciones=observaciones or None,
        usuario_registro=request.user,
        resultados_vigentes=vigentes,
        resultados_vencidos=vencidos,
        resultados_no_encontrados=no_encontrados,
        total_referencias=total
    )

    return JsonResponse({
        "mensaje": "Inspección registrada y comparación realizada con éxito.",
        "inspeccion": {
            "id": inspeccion.id,
            "empresa": inspeccion.empresa,
            "inspector": inspeccion.inspector,
            "codigo": inspeccion.codigo_inspeccion,
            "columna": inspeccion.columna_proveedores,
            "observaciones": inspeccion.observaciones,
            "usuario": request.user.username,
            "fecha": inspeccion.fecha_registro.strftime("%d/%m/%Y"),
            "fecha_inspeccion": fecha_inspeccion,
        },
        "vigentes": vigentes,
        "vencidos": vencidos,
        "no_encontrados": no_encontrados,
        "total_referencia": total,
        "porcentaje_vigentes": f"{round(len(vigentes) * 100 / total, 2)}%" if total else "0%",
        "porcentaje_vencidos": f"{round(len(vencidos) * 100 / total, 2)}%" if total else "0%",
        "porcentaje_no_encontrados": f"{round(len(no_encontrados) * 100 / total, 2)}%" if total else "0%",
        "empleados_por_proveedor": empleados_por_proveedor,
    }, json_dumps_params={"ensure_ascii": False, "indent": 2})'''


    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido. Debe ser POST."}, status=405)

    archivo = request.FILES.get("archivo_excel")
    fecha_inspeccion = request.POST.get("fecha_inspeccion")
    empresa = request.POST.get("empresa", "").strip()
    inspector = request.POST.get("controlador", "").strip()
    codigo = request.POST.get("codigo_inspeccion", "").strip()
    columna_proveedores = request.POST.get("columna_proveedores", "").strip()
    columna_dni = request.POST.get("columna_dni", "").strip()
    observaciones = request.POST.get("observaciones", "").strip()

    errores = []

    # Validación de fecha
    if fecha_inspeccion:
        try:
            fecha_obj = datetime.strptime(fecha_inspeccion, "%Y-%m-%d").date()
        except ValueError:
            errores.append("Formato de fecha_inspeccion inválido, debe ser YYYY-MM-DD.")
            fecha_obj = None
    else:
        fecha_obj = None

    # Validaciones básicas
    if not archivo:
        errores.append("Debe adjuntar un archivo Excel.")
    if not empresa:
        errores.append("Debe ingresar el nombre de la empresa.")
    if not inspector:
        errores.append("Debe ingresar el nombre del inspector.")
    if not columna_proveedores.isdigit():
        errores.append("Debe ingresar un número de columna válido para proveedores.")
    if not columna_dni.isdigit():
        errores.append("Debe ingresar un número de columna válido para DNI.")

    if errores:
        return JsonResponse({"error": "Faltan datos obligatorios.", "detalle": errores}, status=400)

    columna_proveedores = int(columna_proveedores)
    columna_dni = int(columna_dni)

    # Lectura del Excel
    try:
        hojas = pd.read_excel(archivo, engine='openpyxl', sheet_name=None, skiprows=1)
    except Exception as e:
        return JsonResponse({"error": f"Error al leer el Excel: {str(e)}"}, status=400)

    registros = RegistroProveedores.objects.all()
    base = {
        normalizar_nombres(r.nombre_razon_social): {
            "id": r.id,
            "nombre_db": r.nombre_razon_social,
            "fecha_alta": r.fecha_alta,
            "fecha_vto": r.fecha_vto
        }
        for r in registros
    }

    vigentes, vencidos, no_encontrados = [], [], []
    vigentes_set, vencidos_set, no_encontrados_set = set(), set(), set()
    empleados_por_proveedor = {}
    proveedor_dni_set = {}
    total_set = set()

    # Recorremos las hojas del Excel
    for nombre_hoja, df in hojas.items():
        if columna_proveedores >= len(df.columns) or columna_dni >= len(df.columns):
            continue

        mask_valid = df.iloc[:, columna_proveedores].notna() & df.iloc[:, columna_dni].notna()
        proveedores_raw = df.loc[mask_valid, df.columns[columna_proveedores]].astype(str).tolist()
        dnis_raw = df.loc[mask_valid, df.columns[columna_dni]].astype(str).tolist()

        nombres_normalizados = [normalizar_nombres(p) for p in proveedores_raw]
        coincidencias_confirmadas = []
        sin_coincidencia = []

        for i, nombre in enumerate(nombres_normalizados):
            if "empresa" in nombre.lower():
                continue
            total_set.add(nombre)

            match = None
            for base_nombre in base:
                if base_nombre in nombre or nombre in base_nombre:
                    match = base_nombre
                    break

            if match:
                proveedor = base[match]
                proveedor_nombre = proveedor["nombre_db"]

                if proveedor_nombre not in vigentes_set and proveedor_nombre not in vencidos_set:
                    empleados = proveedor_dni_set.get(proveedor_nombre, set())
                    entry = {
                        "referencia": proveedores_raw[i],
                        "coincidencia": proveedor_nombre,
                        "id": proveedor["id"],
                        "fecha_alta": proveedor["fecha_alta"].strftime("%d/%m/%Y") if proveedor["fecha_alta"] else None,
                        "fecha_vto": proveedor["fecha_vto"].strftime("%d/%m/%Y") if proveedor["fecha_vto"] else None,
                        "hoja": nombre_hoja,
                        "empleados": len(empleados)
                    }
                    if (
                        fecha_obj
                        and proveedor["fecha_alta"] is not None
                        and proveedor["fecha_vto"] is not None
                        and proveedor["fecha_alta"] <= fecha_obj <= proveedor["fecha_vto"]
                    ):
                        vigentes.append(entry)
                        vigentes_set.add(proveedor_nombre)
                    else:
                        vencidos.append(entry)
                        vencidos_set.add(proveedor_nombre)

                coincidencias_confirmadas.append((proveedor_nombre, dnis_raw[i]))
            else:
                clave = normalizar_nombres(proveedores_raw[i])
                dni_actual = dnis_raw[i]

                # Si ya existe el proveedor no encontrado, solo agregamos el nuevo DNI si no estaba
                if clave in empleados_por_proveedor:
                    proveedor_dni_set.setdefault(clave, set()).add(dni_actual)
                    empleados_por_proveedor[clave] = len(proveedor_dni_set[clave])
                else:
                    proveedor_dni_set[clave] = {dni_actual}
                    empleados_por_proveedor[clave] = 1

                    # Agregar solo una vez a la lista visible
                    no_encontrados.append({
                        "referencia": proveedores_raw[i],
                        "hoja": nombre_hoja,
                        "empleados": empleados_por_proveedor[clave]
                    })
                    no_encontrados_set.add(clave)

                sin_coincidencia.append((normalizar_nombres(proveedores_raw[i]), dnis_raw[i]))

        # Contar empleados de coincidencias reales
        for proveedor_match, dni in coincidencias_confirmadas:
            if proveedor_match not in empleados_por_proveedor:
                empleados_por_proveedor[proveedor_match] = 0
                proveedor_dni_set[proveedor_match] = set()
            if dni not in proveedor_dni_set[proveedor_match]:
                proveedor_dni_set[proveedor_match].add(dni)
                empleados_por_proveedor[proveedor_match] += 1

    total = len(total_set)

    # Actualizar conteo de empleados en cada item de vigentes y vencidos
    for item in vigentes:
        item["empleados"] = empleados_por_proveedor.get(item["coincidencia"], 0)
    for item in vencidos:
        item["empleados"] = empleados_por_proveedor.get(item["coincidencia"], 0)

    # Crear registro de inspección
    inspeccion = InspeccionProveedores.objects.create(
        empresa=empresa,
        fecha_inspeccion=fecha_inspeccion or None,
        anio=fecha_obj.year if fecha_obj else None,
        inspector=inspector,
        codigo_inspeccion=generar_codigo_unico(),
        columna_proveedores=columna_proveedores,
        columna_dni=columna_dni,
        archivo_excel=archivo,
        observaciones=observaciones or None,
        usuario_registro=request.user,
        resultados_vigentes=vigentes,
        resultados_vencidos=vencidos,
        resultados_no_encontrados=no_encontrados,
        total_referencias=total
    )

    # Ordenar los no encontrados por cantidad de empleados (descendente)
    no_encontrados.sort(key=lambda x: x["empleados"], reverse=True)

    return JsonResponse({
        "mensaje": "Inspección registrada y comparación realizada con éxito.",
        "inspeccion": {
            "id": inspeccion.id,
            "empresa": inspeccion.empresa,
            "inspector": inspeccion.inspector,
            "codigo": inspeccion.codigo_inspeccion,
            "columna": inspeccion.columna_proveedores,
            "observaciones": inspeccion.observaciones,
            "usuario": request.user.username,
            "fecha": inspeccion.fecha_registro.strftime("%d/%m/%Y"),
            "fecha_inspeccion": fecha_inspeccion,
        },
        "vigentes": vigentes,
        "vencidos": vencidos,
        "no_encontrados": no_encontrados,
        "total_referencia": total,
        "porcentaje_vigentes": f"{round(len(vigentes) * 100 / total, 2)}%" if total else "0%",
        "porcentaje_vencidos": f"{round(len(vencidos) * 100 / total, 2)}%" if total else "0%",
        "porcentaje_no_encontrados": f"{round(len(no_encontrados) * 100 / total, 2)}%" if total else "0%",
        "empleados_por_proveedor": empleados_por_proveedor,
    }, json_dumps_params={"ensure_ascii": False, "indent": 2})

def notificar_secretaria(request):
    pdf_file = request.FILES.get('pdf')
    if not pdf_file:
        return JsonResponse({'error': 'No se recibió el archivo PDF.'}, status=400)

    email = EmailMessage(
        subject="Informe de Inspección",
        body="Se adjunta el informe de inspección generado desde el sistema.",
        from_email="notificaciones@sistema.com",
        to=["secretaria@sistema.com"]
    )
    email.attach("Informe_Inspeccion.pdf", pdf_file.read(), "application/pdf")
    email.send()

    return JsonResponse({'success': True})

def listado_inspecciones_view(request):
    inspecciones = InspeccionProveedores.objects.select_related('usuario_registro').order_by('-fecha_registro')
    
    paginator = Paginator(inspecciones, 20)  # 20 por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'inspecciones': page_obj,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
    }

    return render(request, 'proveedores/listado_inspecciones.html', context)

def detalle_inspeccion(request, inspeccion_id):
    inspeccion = get_object_or_404(InspeccionProveedores, id=inspeccion_id)

    
    porc_vigentes = (len(inspeccion.resultados_vigentes)/inspeccion.total_referencias)*100
    porc_vencidos = (len(inspeccion.resultados_vencidos)/inspeccion.total_referencias)*100
    porc_no_encontrados = (len(inspeccion.resultados_no_encontrados)/inspeccion.total_referencias)*100
    context = {
        'inspeccion': inspeccion,
        'vigentes': inspeccion.resultados_vigentes,  
        'vencidos': inspeccion.resultados_vencidos,
        'no_encontrados': inspeccion.resultados_no_encontrados,
        'total_referencias': inspeccion.total_referencias,
        'porc_vigentes' : porc_vigentes ,
        'porc_vencidos' : porc_vencidos ,
        'porc_no_encontrados' : porc_no_encontrados 
        
    }
    return render(request, 'proveedores/detalle_inspeccion.html', context)

@login_required
def proveedores_activos_view(request):
    return render(request,'proveedores/activos.html')

@login_required
def proveedores_view(request):
    concesionarios = Concesionarios.objects.using('catastro').all()
    user = request.user.first_name+' '+request.user.last_name
    if user == ' ':
        user = request.user.username 
    context = {
      'concesionarios' : concesionarios,
      "fecha_actual": datetime.today(),
      "usuario_logueado": user
    }
    return render(request,'proveedores/serch.html',context)

def usuarios_simsa(request):
    # Subconsulta para obtener el nombre de la compañía
    company_name = Companies.objects.using("simsa").filter(
        id=Cast(OuterRef("objectid"), output_field=UUIDField())
    ).values("name")[:1]

    # Subquery para el cuit de la empresa
    company_cuit = Companies.objects.using("simsa").filter(
        id=Cast(OuterRef("objectid"), output_field=UUIDField())
    ).values("cuit")[:1]

    # Consultar usuarios admin de empresas
    usuarios = (
        Users.objects.using("simsa")
        .filter(isdeleted=False)
        .select_related("roleid")
        .annotate(
            company_name=Subquery(company_name),
            company_cuit=Subquery(company_cuit),
        )
        .values("username", "roleid__name", "email", "cuildni", "company_name", "company_cuit")
    )

    # Renderizamos el template pasando los usuarios al contexto
    return render(request, "simsa/usuarios_simsa_list.html", {"usuarios": usuarios})

# ---------------------------
# Usuarios Admin de Empresas
# ---------------------------
def usuarios_pdf(request):
    companies = Companies.objects.using("simsa").filter(isdeleted=False,
        id=Cast(OuterRef("objectid"), output_field=UUIDField())
    ).values("name")[:1]

    usuarios = Users.objects.using("simsa").filter(isdeleted=False,roleid__name='Admin Empresa').select_related("roleid") \
        .annotate(company_name=Subquery(companies)) \
        .values("username", "roleid__name", "email", "company_name")

    # Crear PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    elements.append(Paragraph("Usuarios Admin de Empresas", styles['Title']))

    data = [["#", "Usuario", "Rol", "Email", "Compañía"]] + [
        [i+1, u["username"], u["roleid__name"], u["email"], u.get("company_name", "-")]
        for i, u in enumerate(usuarios)
    ]

    t = Table(data, colWidths=[30, 100, 100, 150, 100])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.darkgrey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey)
    ]))
    elements.append(t)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="usuarios_admin_empresa.pdf"'
    return response


def usuarios_excel(request):
    companies = Companies.objects.using("simsa").filter(isdeleted=False,
        id=Cast(OuterRef("objectid"), output_field=UUIDField())
    ).values("name")[:1]

    usuarios = Users.objects.using("simsa").filter(isdeleted=False,roleid__name='Admin Empresa').select_related("roleid") \
        .annotate(company_name=Subquery(companies)) \
        .values("username", "roleid__name", "email", "company_name")

    df = pd.DataFrame(list(usuarios))
    df.rename(columns={
        "username": "Usuario",
        "roleid__name": "Rol",
        "email": "Email",
        "company_name": "Compañía"
    }, inplace=True)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name="Usuarios Admin Empresa")

    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="usuarios_admin_empresa.xlsx"'
    return response


# ---------------------------
# Empresas sin Usuario
# ---------------------------
def empresas_sin_usuario_excel(request):
    todas_empresas = [
        'ADVANTAGE LITHIUM ARGENTINA S.A.',
'AFRANLLIE JORGE DAVID',
'AGUILAR DE GUAIMAS MABEL ALICIA',
'AGUILAR JOSE ARMANDO',
'AGUILAR JUAN ANGEL',
'ALDEBARAN ARGENTINA S.A.',
'ALEXANDER GOLD GROUP LIMITED',
'ALPHA LITHIUM ARGENTINA S.A.',
'ALPHA MINERALS S.A.U.',
'ALQA LITHIUM S.A.',
'ALTO GRANDE LITHIUM S.A.',
'ALVAREZ CLAUDIA ANAHI',
'ALVAREZ HNOS S.R.L.',
'ANDES MINERAL EXPLORATION S.A.',
'ANGLOGOLD ARGENTINA EXPLORACIONES S.A.',
'ARAYA ALFONSO LUIS',
'ARENAS ALEXIS ADRIEL',
'ARGAÑARAZ OLIVERO FACUNDO PEDRO',
'ARGAÑARAZ OLIVERO RAFAEL',
'ARGENTINA LITIO Y ENERGIA S.A.',
'ARLI S.A.',
'ASTRALI S.A.',
'AVANTI S.R.L.',
'BARR NICOLAS ANDRES',
'BAVIO MIGUEL ALEJANDRO',
'BHP BILLITON WORLD EXPLORATION INC.',
'BOLERA MINERA S.A.',
'BONIFACIO DALMIRO ERNESTO',
'BORAX ARGENTINA S.A.',
'BRINE LITHIUM RESOURCES S.A.',
'BUGANEM CARLOS CRISTIAN',
'BURGOS MIGUEL MARTIN',
'CAIMI FERNANDO ENRIQUE',
'CAMPOS CARLOS ALBERTO',
'CARAPARI S.A.',
'CARDERO ARGENTINA S.A.',
'CARDOZO LUIS EDUARDO',
'CARLOS ALBERTO REPETTO',
'CARRANQUE JORGE DANIEL',
'CASCADERO MINERALS S.A.',
'CASTALIA MINING S.A.',
'CASTILLO HUGO HECTOR',
'CASTILLO ROBERTO JULIO',
'CENTENARIO LITHIUM S.A.',
'CERAMICA ALBERDI S.A.',
'CHILIGUAY FACUNDO ROBERTO',
'COLORADO S.A.',
'COLQUE EXPLORACIONES S.A.',
'COLQUE GERARDO MARTIN',
'COMISION NACIONAL DE ENERGIA ATOMICA',
'CONDORYACU S.R.L.',
'COOPERATIVA DE TRABAJO LA MESA REDONDA LTDA.',
'COOPERATIVA DE TRABAJO MINACLAR LTDA.',
'CORNEJO DIEZ ADOLFO',
'CORNEJO ULADISLAO',
'CORRIENTE ARGENTINA S.A.',
'COSMOS MINERALS S.A.',
'COZZI ELISA ADELA',
'CRITICAL REAGENTS ARGENTINA S.A.S.',
'CRUZ HIPOLITO GUMERCINDO',
'CRUZ JOSE ROBERTO',
'CRUZ LEA ELIANA',
'CRUZ NICOLAS',
'DAGUN ANDREA LORENA',
'DAL BORGO CESAR DOMINGO',
'DAROCA MAURICIO EMILIANO',
'DEFENSA Y ENCAUZAMIENTO S.A.',
'DELGADO PATRICIA NOEMI',
'DEMISA CONSTRUCCIONES S.A.',
'DI GIORGIO OSVALDO',
'DIAZ SERGIO RICARDO',
'DIEGO RUBEN OMAR',
'DINARCO S.A.',
'DOMECQ TRISTAN ALFONSO',
'EGEO S.A.',
'EKEKO S.A.',
'EL PACHAR S.R.L.',
'EL TERRAPLEN S.R.L.',
'ELECTROQUIMICA EL CARMEN S.A.',
'EMISA S.A.',
'ERAMINE SUDAMERICA S.A.',
'ESCALANTE ALBERTO',
'ESCALANTE CARLOS ROBERTO',
'ESPINOSA ALBA ANDREA',
'FERNANDEZ MERCEDES ROSALIA',
'FERNANDEZ PEREZ CARLOS EDUARDO',
'FIGUEROA PATRON VICTOR',
'GALAN LITIO S.A.',
'GANAM MAURELL CARLOS ENRIQUE',
'GAVINOR S.R.L.',
'GEOMIX S.R.L.',
'GEOTERRA S.R.L.',
'GUERRERO JUAN CARLOS',
'GUITIAN RICARDA SALOME',
'GUSTAVO CESAR ANTONIO ASTUDILLO ROSA',
'HANACOLLA S.A.',
'HANAQ ARGENTINA S.A.',
'HANCHA S.A.',
'HANTARA S.A.',
'HASTANA S.A.',
'HERRERA JUAN MANUEL',
'HIJOS DE SALVADOR MUÑOZ S.R.L.',
'HOYOS SIMON AGUSTIN',
'IACUZZI FRANCISCO',
'IGLESIAS LUCAS PEDRO',
'IMERYS MINERALES ARGENTINA S.A.',
'INCOVI S.R.L.',
'INFANTE MARÍA DEL VALLE',
'INGENIERO MEDINA S.A.',
'INTEGRA RECURSOS NATURALES S.A.',
'JUAREZ LEONARDO NICOLAS',
'JUPITER S.R.L.',
'LA CASUALIDAD S.A.',
'LAGOS CLAUDIO ANTONIO',
'LALIN HECTOR DANIEL',
'LAMARCA CARLOS ALBERTO',
'LAS GRAMAS S.A.',
'LIENDRO MARIA ELSA',
'LILAC SOLUTIONS ARIZARO S.A.U.',
'LITHEA INC SUCURSAL ARGENTINA',
'LITHIUM ARGENTINA RESOURCES S.A.',
'LITHIUM S CORPORATION S.A.',
'LITIO ARGENTINO S.A.',
'LITIO MINERA ARGENTINA S.A.',
'LOMA NEGRA COMPAÑIA INDUSTRIAL ARGENTINA S.A.',
'LONDERO JUAN LUIS',
'LOPEZ SABRINA VANESA DEL HUERTO',
'LOPEZ SANCHEZ ROQUE',
'LOZA CALIXTO',
'LOZANO ENRIQUE ADOLFO',
'LOZANO JUAN DOMINGO',
'MAITA SANDRA SUSANA',
'MANSFIELD MINERA S.A.',
'MANUFACTURAS LOS ANDES S.A.',
'MARTINEZ ANGEL GASTON',
'MARTINEZ EDUARDO PATRICIO',
'MARTINEZ OSCAR DAVID',
'MASIE HECTOR',
'MATERIAS PRIMAS ARGENTINA',
'MEDINA HECTOR ENRIQUE',
'M.E.I. OBRAS Y SERVICIOS S.R.L.',
'MELLADO SLEIVE DANIEL EDUARDO',
'MINERA ANSOTANA S.A.',
'MINERA AUSTRAL S.A.',
'MINERA CERRO JUNCAL S.A.',
'MINERA DEL ALTIPLANO S.A.',
'MINERA EL TORO S.A.',
'MINERA INDUSTRIAL ARGENTINA S.R.L.',
'MINERA SANTA RITA S.R.L.',
'MOCOVI S.R.L.',
'MONCHOLI MARIO ANGEL BLAS',
'MONTERRUBIO PEDRO DANIEL',
'MORALES ALEJANDRINA',
'MORALES CAMILO ALBERTO',
'MORALES FELIX HUMBERTO',
'MORALES JUAN',
'MORENO JORGE ENRIQUE',
'MULTIQUIM S.R.L.',
'MUNIAGURRIA CARLOS JORGE',
'MUNICIPALIDAD DE LAS LAJITAS',
'NEVADO MINERALS S.A.',
'NOA ARIZARO S.A.U.',
'NOA LITHIUM BRINES S.A.',
'NOVOA NORMA INES',
'NRG METALS ARGENTINA S.A.',
'NUÑEZ FEDERICO RAMON',
'NUÑEZ RAMON',
'OLIVA GUSTAVO WALTER',
'PACHA MINERALS S.A.',
'PACIFIC RIM MINING CORPORATION ARGENTINA S.A.',
'PADILLA LILIA AURORA',
'PAN AMERICAN ENERGY',
'PESTAÑA DIEGO MARTIN',
'PILCO LEAL PATRICIA OLIVIA',
'PONCE DE LEON ANTONIO ARGENTINO',
'POSCO ARGENTINA S.A.U.',
'POTASIO Y LITIO DE ARGENTINA S.A.',
'POWER MINERALS S.A.',
'PRODUCTORA V-D S.R.L.',
'PROYECTO PASTOS GRANDES S.A.',
'PUCA DE VIVERO SALOME',
'PUNA ARGENTINA S.A.S.',
'PUNA MINING S.A.',
'QUIPILDOR HUGO VICTOR',
'QUIQUINTO ELIZABET AMERICA',
'QUIROZ RENE OSVALDO',
'RAV S.R.L.',
'RECHARGE RESOURCES ARGENTINA S.A.U.',
'RECURSOS ENERGETICOS Y MINEROS SALTA S.A.',
'RENE ELADIO VALDEZ S.R.L.',
'RINCON MINING LIMITED',
'RIODEORO S.A.',
'RIZZOTTI DANIEL ANTONIO',
'RODRIGUEZ ELADIO',
'RODRIGUEZ SILVIA RENE',
'ROJO NATALIA',
'SALTA AMBIENTAL S.R.L.',
'SALTA EXPLORACIONES S.A.',
'SALTA GEOTHERMAL S.A.',
'SANTA INES COPPER S.A.',
'SARAVIA NAVAMUEL HECTOR',
'SCHUBERT JULIO ADOLFO',
'SERVICIOS MINEROS ATACAMA S.R.L.',
'SERVICIOS Y EXPLOTACIONES MINERAS CRUZ S.R.L.',
'SERVICIOS Y PROYECTOS MINEROS S.R.L.',
'SIERRA OSVALDO',
'SILEX ARGENTINA S.A.',
'SOSA ALBERTO RAYMUNDO',
'SOSA QUINTANA RAYMUNDO',
'SOUTHERN EAGLE MINING ARGENTINA S.A.',
'SRUR ANTONIO',
'STUDER SIMON DAVID',
'SULCA SANCHEZ JAVIER FRANCISCO',
'TAMER OSCAR ADOLFO',
'TAPIA JUAN MARCELINO',
'TARITOLAY CESAR ARMANDO',
'TEJERINA ELISA DE SORIANO',
'TOLABA ESCOLASTICO',
'TORTUGA DE ORO S.A.',
'ULEX S.A.',
'ULTRA ARGENTINA S.R.L.',
'VACAREZZA JUAN ESTEBAN ROBUSTIANO',
'VALDEZ JUAN CARLOS',
'VALDEZ NORMA SEBASTIANA',
'VALDEZ RUBEN ELADIO',
'VARGAS ORLANDO SERGIO',
'VENTICOLA GONZALO FEDERICO',
'VIAL ZENTA S.R.L.',
'VIALMANI S.A.',
'VICCO AMALIA MARGARITA',
'VIDAL ENRIQUE JOSE',
'VIENTO BLANCO S.R.L.',
'VIRGILI SAN MILLAN SEBASTIAN',
'VITTONE HECTOR FELIX',
'VITTORI RAUL EDUARDO',
'VIVEROS MARTIN',
'WOMBAT MINERALS S.A.',
'YACONES S.R.L',
'YAMANA ARGENTINA SERVICIOS S.A.',
'YDIARTE ELEUDORO',
'AGUILAR SERGIO IGNACIO',
'TEMPERLEY DIEGO GUILLERMO',
'VIRGILI SAN MILLAN FERNANDO',
'ALVAREZ RODRIGO HORACIO',
'CELORRIO IGNACIO HERNAN',
'MONCHOLI PABLO',
'VARELA JACINTO OSCAR',
'FROMM CARLOS EDUARDO',
'NIOI CLEMENT MILAGROS',
'GOYTIA FERNANDO JOSE',
'PONESSA MARCOS ANTONIO',
'GOMEZ NAAR MARIA',
'NEREO NESTOR MARTIN',
'MASIE WALTER',
'MORALES RICARDO',
'SALAS ALBA SILVIA',
'PADILLA SILVIA DEL ROSARIO',
'SOUTH AMERICAN SALARS S.A.',
'TARITOLAY RESTITUTO',
'BANDI ALBERTO',
'MAURICIO EMILIANO DAROCA S.A.',
'GUERRERO ALVARADO DAVID',
'NIOI CLEMENT FACUNDO',
'PATRON COSTAS FRANCISCO JOSE'
    ]
    companies = Companies.objects.using("simsa").filter(isdeleted=False,
        id=Cast(OuterRef("objectid"), output_field=UUIDField())
    ).values("name")[:1]

    empresas_con_usuario = Users.objects.using("simsa").filter(isdeleted=False) \
        .select_related("roleid") \
        .annotate(company_name=Subquery(companies)) \
        .values_list("company_name", flat=True)

    empresas_sin_usuario = [e for e in todas_empresas if e not in empresas_con_usuario]

    df = pd.DataFrame({"Empresa": empresas_sin_usuario})
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name="Empresas sin Usuario")

    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="empresas_sin_usuario.xlsx"'
    return response

def empresas_sin_usuario_pdf(request):
    # Lista completa de empresas que deberían tener usuario
    todas_empresas = [
        'ADVANTAGE LITHIUM ARGENTINA S.A.',
'AFRANLLIE JORGE DAVID',
'AGUILAR DE GUAIMAS MABEL ALICIA',
'AGUILAR JOSE ARMANDO',
'AGUILAR JUAN ANGEL',
'ALDEBARAN ARGENTINA S.A.',
'ALEXANDER GOLD GROUP LIMITED',
'ALPHA LITHIUM ARGENTINA S.A.',
'ALPHA MINERALS S.A.U.',
'ALQA LITHIUM S.A.',
'ALTO GRANDE LITHIUM S.A.',
'ALVAREZ CLAUDIA ANAHI',
'ALVAREZ HNOS S.R.L.',
'ANDES MINERAL EXPLORATION S.A.',
'ANGLOGOLD ARGENTINA EXPLORACIONES S.A.',
'ARAYA ALFONSO LUIS',
'ARENAS ALEXIS ADRIEL',
'ARGAÑARAZ OLIVERO FACUNDO PEDRO',
'ARGAÑARAZ OLIVERO RAFAEL',
'ARGENTINA LITIO Y ENERGIA S.A.',
'ARLI S.A.',
'ASTRALI S.A.',
'AVANTI S.R.L.',
'BARR NICOLAS ANDRES',
'BAVIO MIGUEL ALEJANDRO',
'BHP BILLITON WORLD EXPLORATION INC.',
'BOLERA MINERA S.A.',
'BONIFACIO DALMIRO ERNESTO',
'BORAX ARGENTINA S.A.',
'BRINE LITHIUM RESOURCES S.A.',
'BUGANEM CARLOS CRISTIAN',
'BURGOS MIGUEL MARTIN',
'CAIMI FERNANDO ENRIQUE',
'CAMPOS CARLOS ALBERTO',
'CARAPARI S.A.',
'CARDERO ARGENTINA S.A.',
'CARDOZO LUIS EDUARDO',
'CARLOS ALBERTO REPETTO',
'CARRANQUE JORGE DANIEL',
'CASCADERO MINERALS S.A.',
'CASTALIA MINING S.A.',
'CASTILLO HUGO HECTOR',
'CASTILLO ROBERTO JULIO',
'CENTENARIO LITHIUM S.A.',
'CERAMICA ALBERDI S.A.',
'CHILIGUAY FACUNDO ROBERTO',
'COLORADO S.A.',
'COLQUE EXPLORACIONES S.A.',
'COLQUE GERARDO MARTIN',
'COMISION NACIONAL DE ENERGIA ATOMICA',
'CONDORYACU S.R.L.',
'COOPERATIVA DE TRABAJO LA MESA REDONDA LTDA.',
'COOPERATIVA DE TRABAJO MINACLAR LTDA.',
'CORNEJO DIEZ ADOLFO',
'CORNEJO ULADISLAO',
'CORRIENTE ARGENTINA S.A.',
'COSMOS MINERALS S.A.',
'COZZI ELISA ADELA',
'CRITICAL REAGENTS ARGENTINA S.A.S.',
'CRUZ HIPOLITO GUMERCINDO',
'CRUZ JOSE ROBERTO',
'CRUZ LEA ELIANA',
'CRUZ NICOLAS',
'DAGUN ANDREA LORENA',
'DAL BORGO CESAR DOMINGO',
'DAROCA MAURICIO EMILIANO',
'DEFENSA Y ENCAUZAMIENTO S.A.',
'DELGADO PATRICIA NOEMI',
'DEMISA CONSTRUCCIONES S.A.',
'DI GIORGIO OSVALDO',
'DIAZ SERGIO RICARDO',
'DIEGO RUBEN OMAR',
'DINARCO S.A.',
'DOMECQ TRISTAN ALFONSO',
'EGEO S.A.',
'EKEKO S.A.',
'EL PACHAR S.R.L.',
'EL TERRAPLEN S.R.L.',
'ELECTROQUIMICA EL CARMEN S.A.',
'EMISA S.A.',
'ERAMINE SUDAMERICA S.A.',
'ESCALANTE ALBERTO',
'ESCALANTE CARLOS ROBERTO',
'ESPINOSA ALBA ANDREA',
'FERNANDEZ MERCEDES ROSALIA',
'FERNANDEZ PEREZ CARLOS EDUARDO',
'FIGUEROA PATRON VICTOR',
'GALAN LITIO S.A.',
'GANAM MAURELL CARLOS ENRIQUE',
'GAVINOR S.R.L.',
'GEOMIX S.R.L.',
'GEOTERRA S.R.L.',
'GUERRERO JUAN CARLOS',
'GUITIAN RICARDA SALOME',
'GUSTAVO CESAR ANTONIO ASTUDILLO ROSA',
'HANACOLLA S.A.',
'HANAQ ARGENTINA S.A.',
'HANCHA S.A.',
'HANTARA S.A.',
'HASTANA S.A.',
'HERRERA JUAN MANUEL',
'HIJOS DE SALVADOR MUÑOZ S.R.L.',
'HOYOS SIMON AGUSTIN',
'IACUZZI FRANCISCO',
'IGLESIAS LUCAS PEDRO',
'IMERYS MINERALES ARGENTINA S.A.',
'INCOVI S.R.L.',
'INFANTE MARÍA DEL VALLE',
'INGENIERO MEDINA S.A.',
'INTEGRA RECURSOS NATURALES S.A.',
'JUAREZ LEONARDO NICOLAS',
'JUPITER S.R.L.',
'LA CASUALIDAD S.A.',
'LAGOS CLAUDIO ANTONIO',
'LALIN HECTOR DANIEL',
'LAMARCA CARLOS ALBERTO',
'LAS GRAMAS S.A.',
'LIENDRO MARIA ELSA',
'LILAC SOLUTIONS ARIZARO S.A.U.',
'LITHEA INC SUCURSAL ARGENTINA',
'LITHIUM ARGENTINA RESOURCES S.A.',
'LITHIUM S CORPORATION S.A.',
'LITIO ARGENTINO S.A.',
'LITIO MINERA ARGENTINA S.A.',
'LOMA NEGRA COMPAÑIA INDUSTRIAL ARGENTINA S.A.',
'LONDERO JUAN LUIS',
'LOPEZ SABRINA VANESA DEL HUERTO',
'LOPEZ SANCHEZ ROQUE',
'LOZA CALIXTO',
'LOZANO ENRIQUE ADOLFO',
'LOZANO JUAN DOMINGO',
'MAITA SANDRA SUSANA',
'MANSFIELD MINERA S.A.',
'MANUFACTURAS LOS ANDES S.A.',
'MARTINEZ ANGEL GASTON',
'MARTINEZ EDUARDO PATRICIO',
'MARTINEZ OSCAR DAVID',
'MASIE HECTOR',
'MATERIAS PRIMAS ARGENTINA',
'MEDINA HECTOR ENRIQUE',
'M.E.I. OBRAS Y SERVICIOS S.R.L.',
'MELLADO SLEIVE DANIEL EDUARDO',
'MINERA ANSOTANA S.A.',
'MINERA AUSTRAL S.A.',
'MINERA CERRO JUNCAL S.A.',
'MINERA DEL ALTIPLANO S.A.',
'MINERA EL TORO S.A.',
'MINERA INDUSTRIAL ARGENTINA S.R.L.',
'MINERA SANTA RITA S.R.L.',
'MOCOVI S.R.L.',
'MONCHOLI MARIO ANGEL BLAS',
'MONTERRUBIO PEDRO DANIEL',
'MORALES ALEJANDRINA',
'MORALES CAMILO ALBERTO',
'MORALES FELIX HUMBERTO',
'MORALES JUAN',
'MORENO JORGE ENRIQUE',
'MULTIQUIM S.R.L.',
'MUNIAGURRIA CARLOS JORGE',
'MUNICIPALIDAD DE LAS LAJITAS',
'NEVADO MINERALS S.A.',
'NOA ARIZARO S.A.U.',
'NOA LITHIUM BRINES S.A.',
'NOVOA NORMA INES',
'NRG METALS ARGENTINA S.A.',
'NUÑEZ FEDERICO RAMON',
'NUÑEZ RAMON',
'OLIVA GUSTAVO WALTER',
'PACHA MINERALS S.A.',
'PACIFIC RIM MINING CORPORATION ARGENTINA S.A.',
'PADILLA LILIA AURORA',
'PAN AMERICAN ENERGY',
'PESTAÑA DIEGO MARTIN',
'PILCO LEAL PATRICIA OLIVIA',
'PONCE DE LEON ANTONIO ARGENTINO',
'POSCO ARGENTINA S.A.U.',
'POTASIO Y LITIO DE ARGENTINA S.A.',
'POWER MINERALS S.A.',
'PRODUCTORA V-D S.R.L.',
'PROYECTO PASTOS GRANDES S.A.',
'PUCA DE VIVERO SALOME',
'PUNA ARGENTINA S.A.S.',
'PUNA MINING S.A.',
'QUIPILDOR HUGO VICTOR',
'QUIQUINTO ELIZABET AMERICA',
'QUIROZ RENE OSVALDO',
'RAV S.R.L.',
'RECHARGE RESOURCES ARGENTINA S.A.U.',
'RECURSOS ENERGETICOS Y MINEROS SALTA S.A.',
'RENE ELADIO VALDEZ S.R.L.',
'RINCON MINING LIMITED',
'RIODEORO S.A.',
'RIZZOTTI DANIEL ANTONIO',
'RODRIGUEZ ELADIO',
'RODRIGUEZ SILVIA RENE',
'ROJO NATALIA',
'SALTA AMBIENTAL S.R.L.',
'SALTA EXPLORACIONES S.A.',
'SALTA GEOTHERMAL S.A.',
'SANTA INES COPPER S.A.',
'SARAVIA NAVAMUEL HECTOR',
'SCHUBERT JULIO ADOLFO',
'SERVICIOS MINEROS ATACAMA S.R.L.',
'SERVICIOS Y EXPLOTACIONES MINERAS CRUZ S.R.L.',
'SERVICIOS Y PROYECTOS MINEROS S.R.L.',
'SIERRA OSVALDO',
'SILEX ARGENTINA S.A.',
'SOSA ALBERTO RAYMUNDO',
'SOSA QUINTANA RAYMUNDO',
'SOUTHERN EAGLE MINING ARGENTINA S.A.',
'SRUR ANTONIO',
'STUDER SIMON DAVID',
'SULCA SANCHEZ JAVIER FRANCISCO',
'TAMER OSCAR ADOLFO',
'TAPIA JUAN MARCELINO',
'TARITOLAY CESAR ARMANDO',
'TEJERINA ELISA DE SORIANO',
'TOLABA ESCOLASTICO',
'TORTUGA DE ORO S.A.',
'ULEX S.A.',
'ULTRA ARGENTINA S.R.L.',
'VACAREZZA JUAN ESTEBAN ROBUSTIANO',
'VALDEZ JUAN CARLOS',
'VALDEZ NORMA SEBASTIANA',
'VALDEZ RUBEN ELADIO',
'VARGAS ORLANDO SERGIO',
'VENTICOLA GONZALO FEDERICO',
'VIAL ZENTA S.R.L.',
'VIALMANI S.A.',
'VICCO AMALIA MARGARITA',
'VIDAL ENRIQUE JOSE',
'VIENTO BLANCO S.R.L.',
'VIRGILI SAN MILLAN SEBASTIAN',
'VITTONE HECTOR FELIX',
'VITTORI RAUL EDUARDO',
'VIVEROS MARTIN',
'WOMBAT MINERALS S.A.',
'YACONES S.R.L',
'YAMANA ARGENTINA SERVICIOS S.A.',
'YDIARTE ELEUDORO',
'AGUILAR SERGIO IGNACIO',
'TEMPERLEY DIEGO GUILLERMO',
'VIRGILI SAN MILLAN FERNANDO',
'ALVAREZ RODRIGO HORACIO',
'CELORRIO IGNACIO HERNAN',
'MONCHOLI PABLO',
'VARELA JACINTO OSCAR',
'FROMM CARLOS EDUARDO',
'NIOI CLEMENT MILAGROS',
'GOYTIA FERNANDO JOSE',
'PONESSA MARCOS ANTONIO',
'GOMEZ NAAR MARIA',
'NEREO NESTOR MARTIN',
'MASIE WALTER',
'MORALES RICARDO',
'SALAS ALBA SILVIA',
'PADILLA SILVIA DEL ROSARIO',
'SOUTH AMERICAN SALARS S.A.',
'TARITOLAY RESTITUTO',
'BANDI ALBERTO',
'MAURICIO EMILIANO DAROCA S.A.',
'GUERRERO ALVARADO DAVID',
'NIOI CLEMENT FACUNDO',
'PATRON COSTAS FRANCISCO JOSE'
    ]
    companies = Companies.objects.using("simsa").filter(isdeleted=False,
        id=Cast(OuterRef("objectid"), output_field=UUIDField())
    ).values("name")[:1]

    # Consultar usuarios admin de empresas
    empresas_con_usuario = list(
        Users.objects.using("simsa")
        .filter(isdeleted=False)
        .select_related("roleid")
        .annotate(company_name=Subquery(companies))
        .values_list("company_name", flat=True)
    )

    # Filtrar empresas que no tienen usuario
    empresas_sin_usuario = [e for e in todas_empresas if e not in empresas_con_usuario]

    # Crear PDF en memoria
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40, leftMargin=40, rightMargin=40)
    elements = []
    styles = getSampleStyleSheet()

    # === Logo arriba a la derecha ===
    logo_path = os.path.join(settings.BASE_DIR, "app/static/img/logo.png")
    if os.path.exists(logo_path):
        img = Image(logo_path, width=200, height=50)  # ajustá tamaño si hace falta
        img.hAlign = "RIGHT"
        elements.append(img)

    elements.append(Spacer(1, 12))

    # === Título ===
    elements.append(Paragraph("Empresas sin Usuario", styles['Heading1']))
    elements.append(Spacer(1, 20))

    # === Tabla ===
    data = [["#", "Empresa"]] + [[i+1, e] for i, e in enumerate(empresas_sin_usuario)]
    t = Table(data, colWidths=[30, 450])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgrey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey])
    ]))
    elements.append(t)

    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="empresas_sin_usuario.pdf"'
    return response

def reporte_empresas_sin_proyecto_pdf(request):
    # Obtener los usuarios Admin Empresa
    users = Users.objects.using("simsa").filter(isdeleted=False,
        roleid__name='Admin Empresa'
    ).values("objectid", "username", "email")

    # Convertir los objectid a UUID
    users_map = {}
    for u in users:
        if u["objectid"]:
            try:
                users_map[uuid.UUID(u["objectid"])] = {
                    "username": u["username"],
                    "email": u["email"]
                }
            except ValueError:
                continue

    users_uuid = list(users_map.keys())

    # Subquery para proyectos activos
    proyectos_activos = Companyprojects.objects.filter(
        companyid=OuterRef('pk'),
        isdeleted=False,
        projectid__isdeleted=False
    )

    # Empresas sin proyecto pero con usuario
    empresas_sin_proyecto = Companies.objects.using("simsa").filter(isdeleted=False,
        id__in=users_uuid
    ).annotate(
        tiene_proyecto=Exists(proyectos_activos)
    ).filter(tiene_proyecto=False)

    # Generar el PDF
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="empresas_sin_proyecto.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # === Logo arriba a la derecha ===
    logo_path = os.path.join(settings.BASE_DIR, "app/static/img/logo.png")
    if os.path.exists(logo_path):
        logo_width = 200
        logo_height = 50
        p.drawImage(
            logo_path,
            width - logo_width - 40,  # 40 px de margen derecho
            height - logo_height - 30,  # 30 px de margen superior
            width=logo_width,
            height=logo_height,
            preserveAspectRatio=True,
            mask="auto"
        )

    # === Título ===
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 60, "Empresas con Usuario sin Proyecto")

    # === Encabezados de la tabla ===
    y = height - 100
    p.setFont("Helvetica-Bold", 12)
    p.drawString(50, y, "Empresa")
    p.drawString(250, y, "Usuario")
    p.drawString(400, y, "Email")
    y -= 20

    # === Contenido ===
    p.setFont("Helvetica", 10)
    for empresa in empresas_sin_proyecto:
        user_info = users_map.get(empresa.id, {"username": "-", "email": "-"})
        p.drawString(50, y, empresa.name[:30])  # truncar por seguridad
        p.drawString(250, y, user_info["username"] or "-")
        p.drawString(400, y, user_info["email"] or "-")
        y -= 15

        # Salto de página
        if y < 50:
            p.showPage()
            y = height - 60

            # Redibujar logo en páginas nuevas
            if os.path.exists(logo_path):
                p.drawImage(
                    logo_path,
                    width - logo_width - 40,
                    height - logo_height - 30,
                    width=logo_width,
                    height=logo_height,
                    preserveAspectRatio=True,
                    mask="auto"
                )

            p.setFont("Helvetica", 10)

    p.save()
    return response

def reporte_empresas_sin_presentacion_pdf(request):
    # Obtener los usuarios Admin Empresa
    users = Users.objects.using("simsa").filter(isdeleted=False,
        roleid__name='Admin Empresa'
    ).values("objectid", "username", "email")

    # Convertir los objectid a UUID y mapear usuario/email
    users_map = {}
    for u in users:
        if u["objectid"]:
            try:
                users_map[uuid.UUID(u["objectid"])] = {
                    "username": u["username"],
                    "email": u["email"]
                }
            except ValueError:
                continue

    users_uuid = list(users_map.keys())

    # Subquery de proyectos activos
    proyectos_activos = Companyprojects.objects.filter(
        companyid=OuterRef('pk'),
        isdeleted=False,
        projectid__isdeleted=False
    )

    # Subquery de presentaciones
    presentaciones = Presentations.objects.filter(
        projectid__companyprojects__companyid=OuterRef('pk'),
        isdeleted=False
    )

    # Empresas que tienen proyecto pero no tienen presentaciones
    empresas_sin_presentacion = Companies.objects.using("simsa").filter(isdeleted=False,
        id__in=users_uuid
    ).annotate(
        tiene_proyecto=Exists(proyectos_activos),
        tiene_presentacion=Exists(presentaciones)
    ).filter(
        tiene_proyecto=True,
        tiene_presentacion=False
    )

    # Generar PDF
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="empresas_sin_presentacion.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4,topMargin=30)
    elements = []
    styles = getSampleStyleSheet()

    # Agregar logo arriba a la derecha
    logo_path = os.path.join(settings.BASE_DIR, "app/static/img/logo.png")  # ajustá el nombre del archivo
    if os.path.exists(logo_path):
        img = Image(logo_path, width=200, height=50)  # tamaño ajustable
        img.hAlign = "RIGHT"
        elements.append(img)

    elements.append(Spacer(1, 12))

    # Título
    elements.append(Paragraph("Empresas con Proyecto sin DDJJ", styles["Heading1"]))
    elements.append(Spacer(1, 20))

    # Tabla
    data = [["Empresa", "Usuario", "Email"]]
    for empresa in empresas_sin_presentacion:
        user_info = users_map.get(empresa.id, {"username": "-", "email": "-"})
        data.append([
            empresa.name,
            user_info["username"] or "-",
            user_info["email"] or "-"
        ])

    table = Table(data, colWidths=[200, 120, 200])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
    ]))

    elements.append(table)
    doc.build(elements)

    return response


def reporte_empresas_sin_proyecto_excel(request):
    # Obtener los usuarios Admin Empresa
    users = Users.objects.using("simsa").filter(isdeleted=False,
        roleid__name='Admin Empresa'
    ).values("objectid", "username", "email")

    # Convertir los objectid a UUID y mapear usuario/email
    users_map = {}
    for u in users:
        if u["objectid"]:
            try:
                users_map[uuid.UUID(u["objectid"])] = {
                    "username": u["username"],
                    "email": u["email"]
                }
            except ValueError:
                continue

    users_uuid = list(users_map.keys())

    # Subquery para proyectos activos
    proyectos_activos = Companyprojects.objects.filter(
        companyid=OuterRef('pk'),
        isdeleted=False,
        projectid__isdeleted=False
    )

    # Empresas sin proyecto pero con usuario
    empresas_sin_proyecto = Companies.objects.using("simsa").filter(isdeleted=False,
        id__in=users_uuid
    ).annotate(
        tiene_proyecto=Exists(proyectos_activos)
    ).filter(tiene_proyecto=False)

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas sin Proyecto"

    # Escribir encabezados
    headers = ["Nombre de Empresa", "Usuario", "Email"]
    ws.append(headers)

    # Escribir datos
    for empresa in empresas_sin_proyecto:
        user_data = users_map.get(empresa.id, {"username": "", "email": ""})
        ws.append([
            empresa.name,
            user_data["username"],
            user_data["email"]
        ])

    # Ajustar ancho de columnas
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 30

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="empresas_sin_proyecto.xlsx"'
    wb.save(response)

    return response

def reporte_empresas_sin_presentacion_excel(request):
    # Obtener los usuarios Admin Empresa
    users = Users.objects.using("simsa").filter(isdeleted=False,
        roleid__name='Admin Empresa'
    ).values("objectid", "username", "email")

    # Convertir los objectid a UUID y mapear usuario/email
    users_map = {}
    for u in users:
        if u["objectid"]:
            try:
                users_map[uuid.UUID(u["objectid"])] = {
                    "username": u["username"],
                    "email": u["email"]
                }
            except ValueError:
                continue

    users_uuid = list(users_map.keys())

    # Subquery de proyectos activos
    proyectos_activos = Companyprojects.objects.filter(
        companyid=OuterRef('pk'),
        isdeleted=False,
        projectid__isdeleted=False
    )

    # Subquery de presentaciones
    presentaciones = Presentations.objects.filter(
        projectid__companyprojects__companyid=OuterRef('pk'),
        isdeleted=False
    )

    # Empresas que TIENEN proyecto pero NO tienen presentaciones
    empresas_sin_presentacion = Companies.objects.using("simsa").filter(
        id__in=users_uuid
    ).annotate(
        tiene_proyecto=Exists(proyectos_activos),
        tiene_presentacion=Exists(presentaciones)
    ).filter(
        tiene_proyecto=True,
        tiene_presentacion=False
    )

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Empresas sin Presentación"

    # Escribir encabezados
    headers = ["Nombre de Empresa", "Usuario", "Email"]
    ws.append(headers)

    # Escribir datos
    for empresa in empresas_sin_presentacion:
        user_data = users_map.get(empresa.id, {"username": "", "email": ""})
        ws.append([
            empresa.name,
            user_data["username"],
            user_data["email"]
        ])

    # Ajustar ancho de columnas
    for i, col in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 30

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="empresas_sin_presentacion.xlsx"'
    wb.save(response)

    return response

def deudas_expedientes(request):
    query = """
    select 
        e."Expediente",
        e."Nombre" as "ExpedienteNombre",
        c2."Name" ,
        case 
            when count(distinct cp."Id") = 3 then '2024 | 2do Semestre'
            when count(distinct cp."Id") = 4 then '2024 | 1er Semestre'
            when count(distinct cp."Id") = 5 then '2023 | 2do Semestre'
            when count(distinct cp."Id") = 6 then '2023 | 1er Semestre'
            else count(distinct cp."Id")::text
        end as semestres_adeudados,
        sum(c."Total") as total_deuda
    from "Canons" c
    inner join "Expedients" e on e."Id" = c."ExpedientId"
    inner join "CanonPeriods" cp on cp."Id" = c."CanonPeriodId"
    inner join "CompanyExpedients" ce on ce."ExpedientId" = e."Id" 
    inner join "Companies" c2 on c2."Id" = ce."CompanyId"
    where c."Paid" = 0
        and e."Tipo" = 'Mina'
        and cp."StartDate" <= CURRENT_DATE
        and ce."IsDeleted" = False
    group by e."Expediente", e."Nombre",c2."Name"
    having count(distinct cp."Id") >= 3
    order by semestres_adeudados desc, total_deuda desc;
    """

    with connections['simsa'].cursor() as cursor:
        cursor.execute(query)
        cols = [col[0] for col in cursor.description]
        results = [dict(zip(cols, row)) for row in cursor.fetchall()]

    return render(request, "simsa/deudas_expedientes.html", {"resultados": results})

def api_proyectos_por_concesionario(request):
    company_id = request.GET.get("company_id")  # lo obtenemos por query param

    if not company_id:
        return JsonResponse({"error": "Falta el parámetro company_id"}, status=400)

    try:
        cps = Companyprojects.objects.using('simsa').filter(companyid=company_id, isdeleted=False)
        proyectos = [
            {
                "id": cp.projectid.id,
                "nombre": cp.projectid.name,
                "anio_creacion": cp.projectid.createdyear,
            }
            for cp in cps
        ]
        return JsonResponse({"proyectos": proyectos})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


def api_periodos_por_proyecto(request):
    print("➡️ Iniciando api_periodos_por_proyecto")

    project_id = request.GET.get("project_id")
    print(f"🟢 Parámetro recibido project_id={project_id}")

    if not project_id:
        print("🔴 Error: falta project_id")
        return JsonResponse({"error": "Falta el parámetro project_id"}, status=400)

    try:
        print("🔍 Consultando presentaciones...")
        presentaciones = (
            Presentations.objects.using('simsa')
            .filter(
                projectid=project_id,
                isdeleted=False,
            )
            .select_related("periodid")
        )

        print(f"📦 Presentaciones encontradas: {presentaciones.count()}")

        # Evitamos duplicados usando un set de periodos
        periodos = []
        vistos = set()

        for p in presentaciones:
            print(f"➡️ Procesando presentación {p.id} | Periodo={getattr(p.periodid, 'name', None)}")
            if p.periodid and p.periodid.name not in vistos:
                periodos.append({
                    "id": p.periodid.id,
                    "nombre": p.periodid.name,
                })
                vistos.add(p.periodid.name)

        print(f"✅ Total de periodos distintos: {len(periodos)}")
        return JsonResponse({"periodos": periodos})

    except Exception as e:
        print("❌ Error en api_periodos_por_proyecto:")
        print(traceback.format_exc())
        return JsonResponse({"error": str(e), "trace": traceback.format_exc()}, status=500)

def generar_informe_proveedores(request):
    concesionario = request.GET.get("concesionario")
    proyecto = request.GET.get("proyecto")
    periodo = request.GET.get("periodo")

    if not (concesionario and proyecto and periodo):
        return JsonResponse({"error": "Faltan parámetros"}, status=400)

    try:
        with connections['simsa'].cursor() as cursor:
            query = f'''
            select 
                c."Cuit",
                c."BusinessName",
                c."RegistrationNumber",
                c."TotalAmount",
                c2."Name" as "Ciiu",
                a."Name" as "Area",
                z."Name" as "Zone",
                nc."Name" as "NativeCommunity",
                cm."Name" as "ContractorMode",
                p2."Name" as "Project",
                c3."Name" as "Company"
            from "Contractors" c
            left join "Ciius" c2 on c2."Id" = c."Id"
            left join "Areas" a on a."Id" = c."AreaId"
            left join "Zones" z on z."Id" = c."ZoneId"
            left join "NativeCommunities" nc on nc."Id" = c."NativeCommunityId"
            left join "ContractorModes" cm on cm."Id" = c."ContractorModeId"
            left join "PresentationContractors" pc on pc."ContractorId" = c."Id"
            left join "Presentations" p on p."Id" = pc."PresentationId"
            left join "Periods" p3 on p."PeriodId" = p3."Id"
            left join "PresentationStates" ps on ps."Id" = p."PresentationStateId"
            left join "Projects" p2 on p."ProjectId" = p2."Id"
            left join "CompanyProjects" cp on cp."ProjectId" = p2."Id"
            left join "Companies" c3 on c3."Id" = cp."CompanyId"
            where p."IsDeleted" = false
              and c3."Id" = %s
              and p2."Id" = %s
              and ps."Name" = 'Presentado'
              and p3."Id" = %s
              and p."IsRectification" = false
              and p2."IsDeleted" = false
              and c3."IsDeleted" = false
              and cp."IsDeleted" = false
            '''
            cursor.execute(query, [concesionario, proyecto, periodo])
            data = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description]

        if not data:
            return JsonResponse({"error": "No se encontraron registros"}, status=404)

        # Crear Excel
        df = pd.DataFrame(data, columns=columns)
        output = io.BytesIO()
        df.to_excel(output, index=False, sheet_name="Informe Proveedores")
        output.seek(0)

        response = HttpResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Informe_Proveedores.xlsx"'
        return response

    except Exception as e:
        import traceback
        return JsonResponse({"error": str(e), "trace": traceback.format_exc()}, status=500)
def reportes_home(request):
    """
    Página principal para la generación de reportes.
    Muestra botones tipo card para cada reporte.

    """

    concesionarios = Companies.objects.using('simsa').all()
    

    context = {
        "concesionarios": concesionarios,
        
        "periodos": ["2023", "2024", "2025"],
        
    }
    return render(request, "simsa/reports.html",context)

def tablero_home(request):
    """
    Página principal para la generación de reportes.
    Muestra botones tipo card para cada reporte.
    """
    return render(request, "simsa/tablero.html")




def consulta_deuda_expediente(request):
    
    return render(request, 'simsa/consulta_expedietes.html')



def consulta_deuda_datos(request):
    """Devuelve solo el expediente buscado."""
    termino = request.GET.get('numero')
    print("Búsqueda:", termino)

    if not termino:
        return HttpResponse("<p style='color:red;'>Debe ingresar un número o nombre de expediente.</p>")

    # Buscar por número o nombre
    expediente = None
    if termino.isdigit():
        expediente = Expedients.objects.using("simsa").filter(expediente=termino, isdeleted=False).first()
    else:
        expediente = Expedients.objects.using("simsa").filter(nombre__icontains=termino, isdeleted=False).first()

    if not expediente:
        return HttpResponse("<p style='color:red;'>No se encontró un expediente con ese número o nombre.</p>")

    concesionarios = (
        expediente.companyexpedients_set
        .filter(isdeleted=False)
        .select_related('companyid')
        .values_list('companyid__name', flat=True)
    )

    pagos = (
        Canons.objects.using("simsa")
        .filter(expedientid=expediente, isdeleted=False)
        .select_related('canonperiodid', 'canonstateid')
        .order_by('-canonperiodid__startdate')
    )

    context = {
        'expediente': expediente,
        'concesionarios': concesionarios,
        'pagos': pagos,
    }

    return render(request, 'simsa/consulta_resultado.html', context)


def expedientes_concesionario(request):
    """Devuelve los demás expedientes del mismo concesionario (con pagos)."""
    expediente_id = request.GET.get('id')

    if not expediente_id:
        return HttpResponse("<p style='color:red;'>Falta el ID del expediente principal.</p>")

    try:
        expediente_principal = Expedients.objects.using("simsa").get(id=expediente_id, isdeleted=False)
    except Expedients.DoesNotExist:
        return HttpResponse("<p style='color:red;'>Expediente no encontrado.</p>")

    # Concesionarios del principal
    concesionarios_ids = expediente_principal.companyexpedients_set.filter(isdeleted=False).values_list("companyid", flat=True)

    # Otros expedientes del mismo concesionario (tipo Mina o Cantera)
    otros_expedientes_qs = (
        Expedients.objects.using("simsa")
        .filter(
            Q(tipo='Mina') | Q(tipo='Cantera'),
            companyexpedients__companyid__in=concesionarios_ids,
            isdeleted=False
        )
        .exclude(pk=expediente_principal.pk)
        .distinct()
    )

    # Construir lista de datos con pagos y concesionarios
    data_expedientes = []
    for exp in otros_expedientes_qs:
        concesionarios = exp.companyexpedients_set.filter(isdeleted=False).select_related('companyid').values_list('companyid__name', flat=True)
        pagos = (
            Canons.objects.using("simsa")
            .filter(expedientid=exp, isdeleted=False)
            .select_related('canonperiodid', 'canonstateid')
            .order_by('-canonperiodid__startdate')
        )
        data_expedientes.append({
            'expediente': exp,
            'concesionarios': concesionarios,
            'pagos': pagos
        })

    context = {"data_expedientes": data_expedientes}
    return render(request, "simsa/otros_expedientes.html", context)

#######################################################PGYPM##########################################################
def pgypm(request):
    return render(request,'pgypm/inspeccion.html')

#################################################EXPEDIENTES##############################################################
def expedientes(request):
    return render(request, 'expedientes/buscar_expediente.html')

################################################SIRGEN######################################################
@login_required
def sirgen_view(request):

    expedientes_queryset = Expediente.objects.filter(expedientedlt=False)

    nro_exp = request.GET.get("nro_exp")
    anio = request.GET.get("anio")
    tipo = request.GET.get("tipo")
    mina = request.GET.get("mina")
    concesionario = request.GET.get("concesionario")

    if nro_exp:
        expedientes_queryset = expedientes_queryset.filter(expedienteid=nro_exp)

    if anio:
        expedientes_queryset = expedientes_queryset.filter(expedienteanio=anio)

    if tipo:
        expedientes_queryset = expedientes_queryset.filter(tipoid=tipo)

    if mina:
        expedientes_queryset = expedientes_queryset.filter(expedientenombremina__icontains=mina)

    if concesionario:
        expedientes_queryset = expedientes_queryset.filter(expedientecaratula__icontains=concesionario)

    # Anotar último pasedestino
    ultimo_pase_destino_subquery = Pase.objects.filter(
    expedienteid=OuterRef('pk')
    ).order_by('-pasenro').values('pasedestino')[:1]

    ultimo_pase_origen_subquery = Pase.objects.filter(
        expedienteid=OuterRef('pk')
    ).order_by('-pasenro').values('paseorigen')[:1]

    ultimo_pase_recibido_subquery = Pase.objects.filter(
        expedienteid=OuterRef('pk')
    ).order_by('-pasenro').values('paserecibido')[:1]

    expedientes_queryset = expedientes_queryset.annotate(
    ultimo_lugar=Subquery(ultimo_pase_destino_subquery),
    ultimo_origen=Subquery(ultimo_pase_origen_subquery),
    ultimo_recibido=Subquery(ultimo_pase_recibido_subquery)
    ).order_by('-expedienteid')

    # Paginación
    paginator = Paginator(expedientes_queryset, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Obtener nombres de organismos
    organismos = {org.organismoid: org.organismonombre for org in Organismo.objects.all()}

    # Agregar nombre del organismo al queryset paginado
    for expediente in page_obj.object_list:
        expediente.ultimo_lugar_nombre = organismos.get(expediente.ultimo_lugar, "Desconocido")
        expediente.ultimo_origen_nombre = organismos.get(expediente.ultimo_origen, "Desconocido")
        expediente.ultimo_esta_recibido = "Sí" if expediente.ultimo_recibido else "No"

    # Armar query string para mantener filtros
    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')
    query_string = urlencode(query_params)

    tipos = Tipo.objects.all()

    context = {
        'now': timezone.now(),
        'page_obj': page_obj,
        'expedientes': page_obj.object_list,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
        'current_page': page_obj.number,
        'total_pages': paginator.num_pages,
        'query_string': query_string,
        'tipo_opciones': tipos,
        'filtros': {
            'nro_exp': nro_exp or '',
            'anio': anio or '',
            'tipo': tipo or '',
            'mina': mina or '',
            'concesionario': concesionario or '',
        }
    }

    return render(request, 'sirgen/serch.html', context)



@login_required
def detalle_expediente(request, expediente_id):
    expediente = get_object_or_404(Expediente, pk=expediente_id)

    # Obtener historial completo y resolver nombres de origen/destino
    historial_raw = Pase.objects.filter(expedienteid=expediente).order_by('-pasenro')
    organismos = {o.organismoid: o.organismonombre for o in Organismo.objects.all()}
    org = Organismo.objects.all()
    for pase in historial_raw:
        pase.pasedestino_nombre = organismos.get(pase.pasedestino, "Desconocido")
        pase.paseorigen_nombre = organismos.get(pase.paseorigen, "Desconocido")

    # PAGINACIÓN
    page_number = request.GET.get('page', 1)
    paginator = Paginator(historial_raw, 10)  # 20 pases por página
    page_obj = paginator.get_page(page_number)

    motivos = Motivo.objects.all()
    lugares = Lugar.objects.all()

    context = {
        'motivos':motivos,
        'lugares':lugares,
        'organismos':org,
        'now': timezone.now(),
        'expediente': expediente,
        'historial_pases': page_obj.object_list,
        'page_obj': page_obj,
        'total_pages': paginator.num_pages,
        'current_page': page_obj.number,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
        'query_string': request.GET.urlencode().replace(f'page={page_number}', '').strip('&'),
    }

    return render(request, 'sirgen/detalle_expediente.html', context)

@login_required
def nuevo_pase(request):
    org = Organismo.objects.all()
    motivos = Motivo.objects.all()
    lugares = Lugar.objects.all()

    context = {
        'motivos':motivos,
        'lugares':lugares,
        'organismos':org,
        'now': timezone.now(),
    }
    return render (request, 'sirgen/new.html',context)

@login_required
def buscar_expediente(request):
    nro = request.GET.get("nro")
    try:
        exp = Expediente.objects.using("catastro").get(expedienteid=nro, expedientedlt=False)
        ultimo_pase = exp.pase_set.order_by('-pasenro').first()

        return JsonResponse({
            "nro": exp.expedienteid,
            "mina": exp.expedientenombremina,
            "anio": exp.expedienteanio,
            "tipo": exp.tipoid.tipo,
            "estado": exp.estadoid.estado,
            "fecha_ultimo_pase": ultimo_pase.pasefecha.strftime("%d/%m/%Y %H:%M") if ultimo_pase else "Sin pase"
        })
    except Expediente.DoesNotExist:
        return JsonResponse({"error": "Expediente no encontrado."})

@login_required
def bandeja_entrada_view(request):

    organismo_id = request.GET.get('organismo_id')

    # Validar que se haya recibido un organismo válido
    if not organismo_id:
        context = {
        'now': timezone.now(),
        'organismos_list': Organismo.objects.all().order_by('organismonombre'),
    }
        return render(request,'sirgen/bandeja_entrada.html',context)

    pases_queryset = Pase.objects.filter(
    pasedestino=organismo_id,
    expedienteid__expedientedlt=False,
    paserecibido=False  # o paserecibido__isnull=True si lo necesitás
    ).select_related('expedienteid').order_by('-pasefecha')


    # Paginación
    paginator = Paginator(pases_queryset, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Obtener nombres de organismos
    organismos = {org.organismoid: org.organismonombre for org in Organismo.objects.all()}

    # Agregar datos auxiliares a cada pase
    for pase in page_obj.object_list:
        pase.organismo_origen_nombre = organismos.get(pase.paseorigen, "Desconocido")
        pase.organismo_destino_nombre = organismos.get(pase.pasedestino, "Desconocido")
    # Armar query string para mantener filtros
    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')
    query_string = urlencode(query_params)

    tipos = Tipo.objects.all()
    
    context = {
        'now': timezone.now(),
        'page_obj': page_obj,
        'pases': page_obj.object_list,
        'has_previous': page_obj.has_previous(),
        'has_next': page_obj.has_next(),
        'previous_page_number': page_obj.previous_page_number() if page_obj.has_previous() else None,
        'next_page_number': page_obj.next_page_number() if page_obj.has_next() else None,
        'current_page': page_obj.number,
        'total_pages': paginator.num_pages,
        'query_string': query_string,
        'tipo_opciones': tipos,
        'organismo_id': organismo_id,
        'organismos_list': Organismo.objects.all().order_by('organismonombre'),
    }


    return render(request,'sirgen/bandeja_entrada.html',context)


@require_POST
@csrf_exempt  # o usá CSRF token con `fetch`
def recibir_pases_masivo(request):
    data = json.loads(request.body)
    pase_ids = data.get('pase_ids', [])

    Pase.objects.filter(id__in=pase_ids, paserecibido=False).update(paserecibido=True)
    return JsonResponse({'status': 'ok'})
@require_POST
def recibir_pase(request, pase_id):
    pase = get_object_or_404(Pase, id=pase_id)
    print(pase)

    
    pase.paserecibido = True
    pase.save()
    messages.success(request, f"Pase #{pase.pasenro} marcado como recibido.")

    # Redirige a donde venía (conserva filtros)
    return redirect(request.META.get('HTTP_REFERER', 'bandeja_entrada'))